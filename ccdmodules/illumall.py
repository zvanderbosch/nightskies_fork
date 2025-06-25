#-----------------------------------------------------------------------------#
#illumall.py
#
#NPS Night Skies Program
#
#Last updated: 2025/05/15
#
#This script computes sky luminance and illuminance
#statistics for all light sources using the skybright
#median filtered mosaic.
#
#Note: 
#
#Input:
#   (1) 
#
#Output:
#   (1) 
#
#History:
#	Zach Vanderbosch -- Created script
#
#-----------------------------------------------------------------------------#

from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

import os
import stat
import arcpy
import numpy as n
import pandas as pd

# Local Source
import filepath
import printcolors as pc

# Set arcpy environment variables
arcpy.env.overwriteOutput = True
arcpy.env.rasterStatistics = "NONE"
arcpy.env.pyramid = "NONE"
arcpy.env.compression = "NONE"
arcpy.CheckOutExtension("Spatial")

# Define print staus prefix
scriptName = 'illumall.py'
PREFIX = f'{pc.GREEN}{scriptName:19s}{pc.END}: '

#------------------------------------------------------------------------------#
#-----------------  Define Fisheye Equal Area Coord System  -------------------#
#------------------------------------------------------------------------------#

geogcs = (
    "GEOGCS["
        "'GCS_Sphere_EMEP',"
        "DATUM['D_Sphere_EMEP',"
        "SPHEROID['Sphere_EMEP',6370000.0,0.0]],"
        "PRIMEM['Greenwich',0.0],"
        "UNIT['Degree',0.0174532925199433]"
    "]"
)
coordinateSystem = (
    "PROJCS["
        "'fisheye equal area',"
        f"{geogcs},"
        "PROJECTION['Lambert_Azimuthal_Equal_Area'],"
        "PARAMETER['False_Easting',0.0],"
        "PARAMETER['False_Northing',0.0],"
        "PARAMETER['Central_Meridian',180.0],"
        "PARAMETER['Latitude_Of_Origin',90.0],"
        "UNIT['Meter',1.0]"
    "]"
)

#------------------------------------------------------------------------------#
#-------------------            Define Functions            -------------------#
#------------------------------------------------------------------------------#

def clear_memory(objectList):
    '''
    Function for clearing variables from memory
    '''
    for obj in objectList:
        if obj:
            del obj


def nl_to_mlux(raster):
    '''
    Unit conversion from nano-Lamberts to milli-Lux
    '''
    return (0.000000761544 * raster) / 314.159


def calc_sky_area_luminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating observed sky luminance is full area
    down to Zenith Angle 96, including below horizon values.
    '''

    # Calculate zonal stats
    outputTable = f"{gridPath}skyhemisall.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaic,outputTable,"DATA","ALL"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results[f'skymax0'] = row.getValue("MAX")
    results[f'skymin0'] = row.getValue("MIN")
    results[f'skyave0'] = row.getValue("MEAN")
    clear_memory([row,rows])

    return results


def calc_sky_mask_luminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating observed all-sky luminance,
    excluding below horizon values.
    '''

    # Calculate zonal stats
    outputTable = f"{gridPath}skyhemismasked.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaic,outputTable,"DATA","ALL"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results[f'skymax1'] = row.getValue("MAX")
    results[f'skymin1'] = row.getValue("MIN")
    results[f'skyave1'] = row.getValue("MEAN")
    clear_memory([row,rows])

    return results


def calc_area_luminouos_emittance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating observed luminous emittance in full
    area down to Zenith Angle 96, including below horizon values.
    '''

    # Convert from nL to mlux units
    mosaicMlux = nl_to_mlux(mosaic)

    # Calculate zonal stats
    outputTable = f"{gridPath}skyhemisall1.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaicMlux,outputTable,"DATA","ALL"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results[f'hemis0'] = row.getValue("MEAN")
    results[f'totalill0'] = row.getValue("SUM")
    clear_memory([row,rows])

    return results


def calc_mask_luminouos_emittance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating observed luminous emittance,
    excluding below-horizon values
    '''

    # Convert from nL to mlux units
    mosaicMlux = nl_to_mlux(mosaic)

    # Calculate zonal stats
    outputTable = f"{gridPath}skyhemisall2.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaicMlux,outputTable,"DATA","ALL"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results[f'hemis1'] = row.getValue("MEAN")
    results[f'totalill1'] = row.getValue("SUM")
    clear_memory([row,rows])

    return results


def calc_scalar_illuminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating observed scalar luminance, using
    all values down to Zenith Angle 90 (full hemisphere).
    '''

    # Convert from nL to mlux units
    mosaicMlux = nl_to_mlux(mosaic)

    # Calculate zonal stats
    outputTable = f"{gridPath}skyscalar.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaicMlux,outputTable,"DATA","SUM"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results['skyscalar'] = row.getValue("SUM")
    clear_memory([row,rows])

    return results


def calc_horizontal_illuminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating horizontal illuminance metrics
    '''

    # Load in horizontal illuminance factor grid
    horizRasterFile = f"{filepath.rasters}horizillf"
    horizRaster = arcpy.sa.Raster(horizRasterFile)

    # Convert from nL to mlux units and multiply by factor grid
    mosaicMlux = nl_to_mlux(mosaic)
    mosaicHoriz = mosaicMlux * horizRaster

    # Calculate zonal stats
    outputTable = f"{gridPath}skyhoriz.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"VALUE",mosaicHoriz,outputTable,"DATA","SUM"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    row = rows.next()
    results['horizs'] = row.getValue("SUM")
    clear_memory([row,rows])

    return results


def calc_vertical_illuminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating vertical illuminance metrics
    '''

    # Convert from nL to mlux units
    mosaicMlux = nl_to_mlux(mosaic)

    # Define array of rotation angles (0 -> 355 in 5-degree increments)
    rotationAngles = n.arange(0,360,5,dtype=int)
    vertIllum = n.zeros(len(rotationAngles))

    # Iterate over rotation angles in 10-degree increments
    vertValues = []
    for i,angle in enumerate(rotationAngles):

        # Status update
        print(f'{PREFIX}Calculating vertical illuminance at angle {angle}...')

        # Load in corresponding raster file
        vertRasterFile = f"{filepath.rasters}vertgrids2/vertf{-angle}"
        vertRaster = arcpy.sa.Raster(vertRasterFile)

        # Calculate vertical illuminance in mlux
        mosaicVert = mosaicMlux * vertRaster

        # Calculate zonal stats
        outputTable = f"{gridPath}skyvert0all.dbf"
        _ = arcpy.sa.ZonalStatisticsAsTable(
            zoneRaster,"VALUE",mosaicVert,outputTable,"DATA","SUM"
        )

        # Extract stats from output table
        rows = arcpy.SearchCursor(outputTable)
        row = rows.next()
        results[f'vert-{angle:03d}'] = row.getValue("SUM")
        vertIllum[i] = row.getValue("SUM")
        clear_memory([row,rows])

    # Get min/max vertical illuminance values and associated azimuth values
    results['max_vlum'] = vertIllum.max()
    results['min_vlum'] = vertIllum.min()
    results['mean_vlum'] = vertIllum.mean()
    results['max_vlum_azimuth'] = rotationAngles[vertIllum.argmax()]
    results['min_vlum_azimuth'] = rotationAngles[vertIllum.argmin()]

    return results


def save_illuminance_data(metrics, dnight, sets, filter):
    '''
    Function to save vertical & horizontal illuminance data to vert.xlsx
    '''

    # Define excel filename and sheet names
    excelFile = f"{filepath.calibdata}{dnight}/vert.xlsx"
    sheet = "Allsky_All_Sources"

    # Delete existing vert.xlsx 
    if os.path.isfile(excelFile):
        os.remove(excelFile)

    # Get number of data sets and define column names
    Nsets = len(sets)
    excelHeaders = [' '] + [f'Set-{i+1}' for i in range(Nsets)]
    excelVertColumns = ['Azimuth'] + ['Vert Illum (mlux)']*Nsets
    excelStatRows = [
        'Min Vert (mlux)',
        'Max Vert (mlux)',
        'Mean Vert (mlux)',
        'Min Vert Azimuth',
        'Max Vert Azimuth',
        'Horiz Illum (mlux)'
    ]

    # Define some cell formatting styles
    headerFont = Font(
        name="Calibri", 
        size=11, 
        bold=True, 
        color="000000"
    )
    headerColor = PatternFill(
        fill_type='solid',
        start_color='CCFFCC', # Green
        end_color='CCFFCC'
    )
    rowColor = PatternFill(
        fill_type='solid',
        start_color='F59067', # Orange
        end_color='F59067'
    )
    headerBorder = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    headerAlignment = Alignment(
        horizontal="center", 
        vertical="center",
        wrap_text=True
    )

    # Add Excel sheets with column headers if it doesn't exist
    with pd.ExcelWriter(excelFile, engine='openpyxl') as writer:

        # Create an openpyxl workbook object
        workbook = writer.book

        # Add sheets and set column/row header formats
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet)
            worksheet = writer.sheets[sheet]
            for i, value in enumerate(excelHeaders,1):
                cell = worksheet.cell(row=1, column=i)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = headerColor
                cell.alignment = headerAlignment
                colLetter = get_column_letter(i)
                worksheet.column_dimensions[colLetter].width = 12
            for i, value in enumerate(excelStatRows,1):
                cell = worksheet.cell(row=i+1, column=1)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = rowColor
                cell.alignment = headerAlignment
            for i, value in enumerate(excelVertColumns,1):
                cell = worksheet.cell(row=10, column=i)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = headerColor
                cell.alignment = headerAlignment
            
            # Set width of first column separately
            worksheet.column_dimensions['A'].width = 20

        # Grab the worksheet
        worksheet = writer.sheets[sheet]

        # Iterate over each data set
        for s in sets:

            # Convert set number to integer
            setnum = int(s[0])

            # Get metrics associated with the set/filter
            setIndex = (
                (metrics['dataset'] == setnum) &
                (metrics['filter'] == filter)
            )
            setMetrics = metrics.loc[setIndex]

            # Add statistics values
            worksheet.cell(row=2, column=setnum+1, value=setMetrics['min_vlum'].iloc[0])
            worksheet.cell(row=3, column=setnum+1, value=setMetrics['max_vlum'].iloc[0])
            worksheet.cell(row=4, column=setnum+1, value=setMetrics['mean_vlum'].iloc[0])
            worksheet.cell(row=5, column=setnum+1, value=setMetrics['min_vlum_azimuth'].iloc[0])
            worksheet.cell(row=6, column=setnum+1, value=setMetrics['max_vlum_azimuth'].iloc[0])
            worksheet.cell(row=7, column=setnum+1, value=setMetrics['horizs'].iloc[0])
            worksheet.cell(row=2, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=3, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=4, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=5, column=setnum+1).number_format = '0'
            worksheet.cell(row=6, column=setnum+1).number_format = '0'
            worksheet.cell(row=7, column=setnum+1).number_format = '0.00000'

            # Get vertical illuminance columns for the given sheet
            vertCols = [c for c in metrics.columns if 'vert-' in c]

            # Get azimuth and vertical illuminance
            vertValues = setMetrics[vertCols].values[0]
            azValues = [int(x.split("-")[1]) for x in vertCols]

            # Add azimuth and vertical illuminance values to table
            for j,(az,vert) in enumerate(zip(azValues,vertValues)):
                
                # Add azimuth values for first set only
                if setnum == 1:
                    worksheet.cell(row=j+11, column=1, value=az)
                    worksheet.cell(row=j+11, column=1).number_format = '0'

                # Add vertical illuminance values
                worksheet.cell(row=j+11, column=setnum+1, value=vert)
                worksheet.cell(row=j+11, column=setnum+1).number_format = '0.00000'


#------------------------------------------------------------------------------#
#-------------------              Main Program              -------------------#
#------------------------------------------------------------------------------#

def calculate_statistics(dnight,sets,filter):
    '''
    Main program for computing sky luminance and illuminance
    statistics from only anthropogenic sources. 
    '''

    # Filter paths
    F = {'V':'', 'B':'B/'}

    # Set ArcGIS working directories
    scratchsetp = f"{filepath.rasters}scratch_metrics/"
    arcpy.env.workspace = scratchsetp
    arcpy.env.scratchWorkspace = scratchsetp

    # Load in mask rasters for zonal stats
    maskRaster = arcpy.sa.Raster(f"{filepath.griddata}{dnight}/mask/maskd.tif")
    areaRaster = arcpy.sa.Raster(f"{filepath.rasters}arearasterf")
    hemiRaster = arcpy.sa.Raster(f"{filepath.rasters}hemirasterf")

    # Loop through each data set
    illumallOutput = []
    for s in sets:

        # Set path for grid datasets
        setnum = int(s[0])
        gridsetp = f"{filepath.griddata}{dnight}/S_{setnum:02d}/{F[filter]}"

        # Initial status update for data set
        print(f'{PREFIX}Processing {dnight} Set-{setnum} {filter}-band...')

        # Load in median sky brightness mosaic in nano-Lamberts [nL]
        inRaster = arcpy.sa.Raster(f"{gridsetp}median/skybrightnl")

        # Project raster into fisheye coordinate system
        brightRasterFile = "skynlf.tif"
        arcpy.ProjectRaster_management (
            inRaster, brightRasterFile, coordinateSystem, "BILINEAR","5558.8"
        )
        brightRaster = arcpy.sa.Raster(brightRasterFile)

        # Initialize result dict
        resultDict = {}

        # Calculate sky luminance metrics
        print(f'{PREFIX}Calculating sky luminance for all sources...')
        resultDict = calc_sky_area_luminance(brightRaster, areaRaster, gridsetp, resultDict)
        resultDict = calc_sky_mask_luminance(brightRaster, maskRaster, gridsetp, resultDict)


        # Calculate luminous emittance
        print(f'{PREFIX}Calculating luminous emittance for all sources...')
        resultDict = calc_area_luminouos_emittance(brightRaster, areaRaster, gridsetp, resultDict)
        resultDict = calc_mask_luminouos_emittance(brightRaster, maskRaster, gridsetp, resultDict)


        # Calculate scalar illuminance
        print(f'{PREFIX}Calculating scalar illuminance for all sources...')
        resultDict = calc_scalar_illuminance(brightRaster, hemiRaster, gridsetp, resultDict)


        # Calculate horizontal illuminance
        print(f'{PREFIX}Calculating horizontal illuminance for all sources...')
        resultDict = calc_horizontal_illuminance(brightRaster, hemiRaster, gridsetp, resultDict)


        # Calculate vertical illuminance
        print(f'{PREFIX}Calculating vertical illuminance for all sources...')
        resultDict = calc_vertical_illuminance(brightRaster, areaRaster, gridsetp, resultDict)

        # Generate dataframe entry for given dataset
        illumallEntry = pd.DataFrame(
            {
                'datanight': dnight,
                'dataset': setnum,
                'filter': filter,
                **resultDict
            },
            index = [setnum-1]
        )
        illumallOutput.append(illumallEntry)


    # Create final dataframe output
    illumallOutput = pd.concat(illumallOutput)

    # Save illuminance data to vert.xlsx file
    save_illuminance_data(illumallOutput, dnight, sets, filter)

    return illumallOutput
