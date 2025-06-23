#-----------------------------------------------------------------------------#
#savetables.py
#
#NPS Night Skies Program
#
#Last updated: 2025/05/28
#
#This script generates the final output tables consolidating
#data night attributes, calibration, and light pollution 
#metrics for each data set.
#
#Note: 
#
#Input:
#   (1) 
#
#Output:
#   (1) 
#
#History:
#	Zach Vanderbosch -- Created script
#
#-----------------------------------------------------------------------------#

from datetime import datetime
from astropy.io import fits
from astropy.time import Time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

import numpy as n
import pandas as pd
import astropy.units as u

# Local Source
import filepath
import printcolors as pc

# Define print status prefix
scriptName = 'savetables.py'
PREFIX = f'{pc.GREEN}{scriptName:19s}{pc.END}: '


#------------------------------------------------------------------------------#
#-------------------         Define Global Variables        -------------------#
#------------------------------------------------------------------------------#

# Titles, column names, and columns widths for each sheet
SHEETDATA = {
    'NIGHT METADATA':{
        'title':"DATA NIGHT ATTRIBUTES",
        'colNames':[
            'DNIGHT', 'NPS_UNIT', 'UNIT_CODE', 'LONGITUDE', 'LATITUDE', 
            'ELEVATION', 'SITE_NAME', 'DATE_START_UT', 'TIME _START_UT', 
            'AIR_TEMP_C', 'RH', 'WIND_SPEED_MPH', 'CAMERA', 'CAMERA_TEMP', 
            'LENS', 'FILTER', 'INSTRUMENT', 'NUM_IMAGES', 'EXPTIME', 'ZEROPOINT', 
            'IMG_SCALE_OFFSET', 'NUM_SETS', 'ZLM', 'BORTLE', 'ALBEDO', 'SQM', 
            'OBS_1', 'OBS_2', 'OBS_3', 'OBS_4', 'NARRATIVE', 'PICTAZIMUTH'
        ],
        'colWidths':[
            16.9, 36.0, 23.1, 15.9, 13.1, 14.9, 43.6, 18.7, 18.6, 17.3,  8.0, 
            17.9, 12.4, 16.0,  8.7, 10.1, 17.0, 16.4,  9.6, 14.3, 20.7, 13.6, 
            10.3, 15.0, 13.0,  9.6, 23.0, 22.3, 21.0, 23.4,176.3, 33.9
        ]
    },
    'CITIES':{
        'title':"NEARBY 100 CITIES BEARING AND DISTANCE (KM) ORDERED ACCORDING TO WALKER'S LAW",
        'colNames':[
            'DNIGHT', 'PLACE', 'STATE', 'WALKERS', 'POPULATION', 'DISTANCE', 
            'BEARING', 'HALF_WIDTH_DEG'
        ],
        'colWidths':[
            15.4, 37.0, 13.0, 11.9, 15.7, 11.7, 10.9, 20.9
        ]
    },
    'SET METADATA':{
        'title':"DATA SET ATTRIBUTES",
        'colNames':[
            'DNIGHT', 'DSET', 'SSTART_DATE_UT', 'SSTART_TIME_UT', 'MID_DATE_LMT', 
            'MID_TIME_LMT', 'GLARE', 'ATMOSPHERE', 'COLLECTION', 'PROCESSING', 
            'REFERENCE', 'USEABLE', 'CLOUDS', 'PLUMES', 'PCT20', 'COLLECTION_NOTES'
        ],
        'colWidths':[
            14.1, 13.0, 17.9, 18.1, 16.0, 15.6, 10.0, 14.6, 
            12.1, 13.0, 12.0, 11.4, 12.7, 13.0, 13.0, 56.7
        ]
    },
    'CALIBRATION':{
        'title':"IMAGE CALIBRATION INFORMATION",
        'colNames':[
            'DNIGHT', 'DSET', 'FLAT_NAME', 'CURVE_NAME', 'BIAS_DRIFT', 
            'MAX_POINT_ERR', 'AVE_POINT_ERR', 'BAD_FRAMES', 'CALIB_NOTES'
        ],
        'colWidths':[
            14.7, 13.0, 36.3, 13.7, 16.6, 18.0, 17.0, 16.1, 105.3
        ]
    },
    'EXTINCTION':{
        'title':"ALL SKY EXTINCTION REGRESSION INFORMATION",
        'colNames':[
            'DNIGHT', 'DSET', 'NUM_SOLVED', 'STARS_REG', 'STARS_REJ', 'STD_ERR_Y', 
            'ZEROPOINT_FLOAT', 'COLOR_FLOAT', 'ZEROPOINT', 'COLOR', 'EXTCOEFF'
        ],
        'colWidths':[
            16.3, 13.0, 17.9, 17.6, 15.0, 16.7, 20.0, 16.0, 15.4, 13.0, 15.3
        ]
    },
    'IMG COORDS':{
        'title':"COORDINATES OF CENTER OF EACH IMAGE",
        'colNames':[
            'DNIGHT', 'DSET', 'IMG_NUM', 'AZIMUTH', 'ALTITUDE', 'RA', 'DEC', 
            'GALL', 'GALB', 'ECL', 'ECB'
        ],
        'colWidths':[
            15.0, 9.1, 11.1, 10.4, 10.9, 9.1, 13.0, 13.0, 13.0, 13.0, 13.0
        ]
    },
    'V2 PHOTOMETRY':{
        'title':"ALL SKY PHOTOMETRY STATISTICS FROM VERSION 2 SPREADSHEET METHODS",
        'colNames':[
            'DNIGHT', 'DSET', 'ZENITH_MSA_V2', 'ALLSKY_MAGS_V2', 'ZA70_MAGS_V2', 
            'BRIGHTEST_MSA_V2', 'DARKEST_MSA_V2', 'SYN_SQM', 'SCALAR_ILL_V2'
        ],
        'colWidths':[
            16.0, 13.0, 17.7, 19.7, 17.9, 19.4, 18.3, 15.0, 14.0
        ]
    },
    'V4 PHOTOMETRY':{
        'title':"ALL SKY PHOTOMETRY STATISTICS FROM VERSION 4 GRID PROCESSING METHOD",
        'colNames':[
            'DNIGHT', 'DSET', 'AVE_LUM_MSA', 'AVE_LUM_MCCD', 'ZENITH_LUM_MSA', 
            'ZENITH_LUM_MCCD', 'BRIGHTEST_LUM_MSA', 'BRIGHTEST_LUM_MCCD', 
            'ALLSKY_MAGS', 'ALLSKY_MLX', 'HORIZ_MLX', 'MAXVERT_MLX', 'NSTARS_FLAT', 
            'NSTARS_OBS', 'NSTARS_EXT', 'VISSTARS_NAT', 'VISSTARS_OBS', 
            'PCT_VISSTARS', 'SCALAR_ILL_V4'
        ],
        'colWidths':[
            16.3, 13.0, 20.0, 16.6, 20.6, 21.1, 22.4, 24.0, 19.4, 
            15.1, 15.6, 17.1, 15.6, 14.9, 15.1, 15.3, 15.1, 15.1, 16.6
        ]
    },
    'GLARE':{
        'title':"PHOTOMETRY OF GLARE SOURCES BY CCD CAMERA WITH ND FILTER",
        'colNames':[
            'DNIGHT', 'DSET', 'ALLSKY_GLARE_MAGS', 'ALLSKY_GLARE_MLX', 
            'MAXVERT_GLARE_MLX', 'MEANVERT_GLARE_MLX', 'MINVERT_GLARE_MLX', 
            'HORIZ_GLARE_MLX'
        ],
        'colWidths':[
            16.1, 13.0, 20.9, 20.4, 24.3, 27.6, 22.9, 20.1
        ]
    },
    'NATSKY':{
        'title':"NATURAL SKY MODEL FIT INFORMATION",
        'colNames':[
            'DNIGHT', 'DSET', 'EM_LYR_HT_KM', 'SITE_ELEV', 'EXTCOEFF_USED', 
            'ZENITH_AIRGLOW_MCCD', 'AIRGLOW_EXT_CONST', 'ADL_MULT', 'ZOD_EXT_CONST', 
            'GAL_EXT_CONST', 'FIT_QUALITY', 'NATSKY_FIT_NOTES'
        ],
        'colWidths':[
            15.1, 10.0, 17.7, 12.4, 17.1, 25.3, 
            22.7, 13.0, 16.6, 18.1, 14.9, 80.9
        ]
    },
    'LP':{
        'title':"LIGHT POLLUTION CALCULATIONS FROM SKYGOW (LUMINANCE AND ILLUMINANCE) FOR THE WHOLE OBSERVABLE SKY",
        'colNames':[
            'DNIGHT', 'DSET', 'SQI_ALLSKY', 'ALR', 'ALLSKY_ART_MAGS', 
            'ALLSKY_ART_MLX', 'MAXVERT_ART_MLX', 'MAXVERT_LPR', 
            'MEANVERT_ART_MLX', 'MEANVERT_LPR', 'MINVERT_ART_MLX', 'MINVERT_LPR', 
            'HORIZ_ART_MLX', 'HORIZ_LPR', 'BRIGHTEST_ART_MCCD', 'BRIGHTEST_LPR', 
            'ZENITH_LUM_ART_MCCD', 'ZENITH_LUM_LPR', 'MEANLUM_ART_MCCD', 'MODEL_ALR'
        ],
        'colWidths':[
            14.3, 13.0, 13.1,  9.7, 24.6, 23.1, 20.9, 20.1, 21.3, 17.1, 
            21.6, 16.6, 18.4, 15.3, 25.3, 18.1, 25.0, 18.7, 23.3, 15.1
        ]
    },
    'LP80':{
        'title':"LIGHT POLLUTION CALCULATIONS FROM SKYGOW (LUMINANCE AND ILLUMINANCE) FOR SKY FROM ZENITH TO ZENITH ANGLE 80",
        'colNames':[
            'DNIGHT', 'DSET', 'SQI_ZA80', 'ZA80_ALR', 'ZA80_ART_MAGS', 
            'ZA80_ART_MLX', 'ZA80_MAXVERT_ART_MLX', 'ZA80_MAXVERT_LPR', 
            'ZA80_ MEANVERT_ART_MLX', 'ZA80_MEANVERT_LPR', 'ZA80_MINVERT_ART_MLX', 
            'ZA80_MINVERT_LPR', 'ZA80_HORIZ_ART_MLX', 'ZA80_HORIZ_LPR', 
            'ZA80_BRIGHTEST_ART_MCCD', 'ZA80_BRIGHTEST_LPR', 'ZA80_MEANLUM_ART_MCCD'
        ],
        'colWidths':[
            12.4, 13.0, 13.1,  9.7, 24.6, 23.1, 26.0, 20.1, 26.3, 
            22.6, 25.9, 21.3, 22.6, 20.4, 28.6, 22.4, 27.7
        ]
    },
    'LP70':{
        'title':"LIGHT POLLUTION CALCULATIONS FROM SKYGOW (LUMINANCE AND ILLUMINANCE) FOR SKY FROM ZENITH TO ZENITH ANGLE 70",
        'colNames':[
            'DNIGHT', 'DSET', 'SQI_ZA70', 'ZA70_ALR', 'ZA70_ART_MAGS', 
            'ZA70_ART_MLX', 'ZA70_MAXVERT_ART_MLX', 'ZA70_MAXVERT_LPR', 
            'ZA70_ MEANVERT_ART_MLX', 'ZA70_MEANVERT_LPR', 'ZA70_MINVERT_ART_MLX', 
            'ZA70_MINVERT_LPR', 'ZA70_HORIZ_ART_MLX', 'ZA70_HORIZ_LPR', 
            'ZA70_BRIGHTEST_ART_MCCD', 'ZA70_BRIGHTEST_LPR', 'ZA70_MEANLUM_ART_MCCD'
        ],
        'colWidths':[
            12.4, 13.0, 13.1,  9.7, 24.6, 23.1, 26.0, 20.1, 26.3, 
            22.6, 25.9, 21.3, 22.6, 20.4, 28.6, 22.4, 27.7
        ]
    },
    'ZONES':{
        'title':"LIGHT POLLUTION FROM SKYGLOW BY ZENITH ANGLE ZONE",
        'colNames':[
            'DNIGHT', 'DSET', 'ZONE1_MCCD', 'ZONE1_MLX', 'ZONE1_PCT', 'ZONE2_MCCD', 
            'ZONE2_MLX', 'ZONE2_PCT', 'ZONE3_MCCD', 'ZONE3_MLX', 'ZONE3_PCT', 
            'ZONE4_MCCD', 'ZONE4_MLX', 'ZONE4_PCT', 'ZONE5_MCCD', 'ZONE5_MLX', 
            'ZONE5_PCT'
        ],
        'colWidths':[
            16.0, 13.0, 16.6, 13.7, 14.0, 14.6, 15.4, 16.1, 16.9, 
            17.3, 17.1, 17.7, 17.0, 13.7, 14.9, 16.7, 12.9
        ]
    },
    'V4 PERCENTILES ALL':{
        'title':"PERCENTILE STATISTICS HORIZON MASKED ALL SOURCES",
        'colNames':[
            'DNIGHT', 'DSET', 'P05DEG_ALL', 'P1DEG_ALL', 'P99_ALL', 'P98_ALL', 
            'P95_ALL', 'P90_ALL', 'P80_ALL', 'P70_ALL', 'P60_ALL', 'P50_ALL', 
            'P01_ALL', 'P0005_ALL'
        ],
        'colWidths':[
            18.6, 9.1, 14.3, 13.0,  9.1, 13.0, 13.0, 
            13.0, 9.4,  9.1, 13.0, 11.0, 13.3, 16.1
        ]
    },
    'V4 PERCENTILES LP':{
        'title':"PERCENTILE STATISTICS HORIZON MASKED ARTIFICIAL ONLY",
        'colNames':[
            'DNIGHT', 'DSET', 'P05DEG_ART', 'P1DEG_ART', 'P99_ART', 'P98_ART', 
            'P95_ART', 'P90_ART', 'P80_ART', 'P70_ART', 'P60_ART', 'P50_ART', 
            'P01_ART'
        ],
        'colWidths':[
            18.6, 9.1, 14.3, 13.0, 9.1, 13.0, 13.0, 
             9.0, 9.7,  9.1, 13.0, 9.4, 11.3
        ]
    }
}


# Style definitions applied to various cell types
SHEETSTYLES = {
    'main_title': {
        'font': Font(
            name="Calibri", 
            size=16, 
            bold=True, 
            color="000000"
        )
    },
    'nightdata_headers': {
        'font': Font(
            name="Calibri", 
            size=11, 
            bold=True, 
            color="000000"
        )
    },
    'nightdata_entries': {
        'font': Font(
            name="Calibri", 
            size=10, 
            bold=True, 
            color="000000"
        ),
        'alignment_lb': Alignment(
            horizontal="left", 
            vertical="bottom",
            wrap_text=False
        ),
        'alignment_rb': Alignment(
            horizontal="right", 
            vertical="bottom",
            wrap_text=False
        )
    },
    'sheet_titles': {
        'font': Font(
            name="Calibri", 
            size=11, 
            bold=False, 
            color="000000"
        )
    },
    'column_headers': {
        'font': Font(
            name="Calibri", 
            size=11, 
            bold=True, 
            color="000000"
        ),
        'border': Border(
            bottom=Side(style='thin')
        ),
        'alignment_cb': Alignment(
            horizontal="center", 
            vertical="bottom",
            wrap_text=False
        ),
        'alignment_lb': Alignment(
            horizontal="left", 
            vertical="bottom",
            wrap_text=False
        )
    },
    'data_fields': {
        'font': Font(
            name="Courier", 
            size=10, 
            bold=False, 
            color="000000"
        ),
        'alignment_cb': Alignment(
            horizontal="center", 
            vertical="bottom",
            wrap_text=False
        ),
        'alignment_lb': Alignment(
            horizontal="left", 
            vertical="bottom",
            wrap_text=False
        )
    },
    'narrative': {
        'font': Font(
            name="Calibri", 
            size=11, 
            bold=False, 
            color="000000"
        ),
        'alignment': Alignment(
            horizontal="left", 
            vertical="bottom",
            wrap_text=True
        )
    },
    'datetime':{
        'datestyle': NamedStyle(
            name='date_style', 
            number_format='d-mmm-yy'
        ),
        'timestyle': NamedStyle(
            name='time_style', 
            number_format='h:mm:ss'
        )
    }
}


# Dropdown menu options for some NIGHT METADATA cells
DROPDOWNS = {
    'CAMERA': [
        'IMG1','ML 1','ML 2','ML 3','ML 4',
        'ML 5','ML 6','ML 7', 'SBIG 1','Proline'
    ],
    'LENS': [
        'Nikon 1.2', 'Nikon 1.8'
    ],
    'FILTER': [
        '5541','5728','6084','7706','8582','8525','9047','9776'
    ],
    'INSTRUMENT': [
        'NexStar 1','NexStar 2','NexStar 3','NexStar 4','NexStar 5',
        'NexStar 6','NexStar 7','NexStar 8','NexStar 9','NexStar 10',
        'NexStar 11','NexStar SE1'
    ],
    'OBSERVERS': [
        'B Banet',
        'D Duriscoe',
        'S Hummel',
        'L Hung',
        'T Jiles',
        'K Magargal',
        'B Meadows',
        'B Miller',
        'C Moore',
        'A Pipkin',
        'Z Vanderbosch',
        'J White'
    ]
}


# Some mappings between FITS header values and standard NIGHT METADATA values
METADATA_MAPPINGS = {
    'INSTRUMENT': {
        'Nexstar 1': 'NexStar 1',
        'Nexstar 2': 'NexStar 2',
        'Nexstar 3': 'NexStar 3',
        'Nexstar 4': 'NexStar 4',
        'Nexstar 5': 'NexStar 5',
        'Nexstar 6': 'NexStar 6',
        'Nexstar 7': 'NexStar 7',
        'Nexstar 8': 'NexStar 8',
        'Nexstar 9': 'NexStar 9',
        'Nexstar 10': 'NexStar 10',
        'Nexstar 11': 'NexStar 11',
        'Nexstar SE1': 'NexStar SE1'
    },
    'CAMERA': {
        'ML1': 'ML 1',
        'ML1': 'ML 2',
        'ML3': 'ML 3',
        'ML4': 'ML 4',
        'ML5': 'ML 5',
        'ML6': 'ML 6',
        'ML7': 'ML 7'
    },
    'LENS': {
        'Nikon f/2_new': 'Nikon 1.2'
    },
    'FILTER': {
        'V': {
            'ML 2': 9776,
            'ML 3': 9047,
            'ML 4': 9776,
            'ML 5': 9776
        },
        'B': {}
    }

}
#------------------------------------------------------------------------------#
#-------------------            Define Functions            -------------------#
#------------------------------------------------------------------------------#

def nl_to_mccd(x):
    '''
    Convert nano-Lambert to micro-Candela units
    '''
    return 3.1831 * x


def nl_to_mags(x):
    '''
    Convert nano-Lambert to mag/arcsecond^2
    '''
    return 26.3308 - 2.5*n.log10(x)


def nl_to_mlux(raster):
    '''
    Unit conversion from nano-Lamberts to milli-Lux
    '''
    return (0.000000761544 * raster) / 314.159


def mags_to_mccd(x):
    '''
    Convert mag/arcsecond^2 to micro-Candela
    '''
    return 3.1831 * (34.08 * n.exp(20.7233 - 0.92104*x))


def mlux_to_mag(x):
    '''
    Convert milli-lux to magnitudes
    '''
    return -13.98 - 2.5*n.log10(x/1000)


def mlux_to_nl(x):
    '''
    Convert milli-Lux to nano-Lambert units
    '''
    return 412529264 * x


def create_excel_template(excelFile):
    '''
    Function that creates the Excel template for storing output tables.
    '''

    # Initial creation and formatting of excel file
    with pd.ExcelWriter(excelFile, engine='openpyxl', mode='w') as writer:

        # Create an openpyxl workbook object
        workbook = writer.book

        # Create and format each sheet
        for sheetname in SHEETDATA.keys():

            # Create sheet
            workbook.create_sheet(sheetname)
            worksheet = writer.sheets[sheetname]

            # Add sheet titles
            if sheetname == "NIGHT METADATA":

                # Add primary title for entire workbook
                cell = worksheet.cell(row=1, column=1)
                cell.value = "NPS Night Skies Program All Sky Imaging Data Report"
                cell.font = SHEETSTYLES['main_title']['font']

                # Add 'processed by' and 'date generated' fields
                cell = worksheet.cell(row=1, column=7)
                cell.value = "Processed by:"
                cell.font = SHEETSTYLES['nightdata_headers']['font']
                cell = worksheet.cell(row=1, column=8)
                cell.value = "Date generated:"
                cell.font = SHEETSTYLES['nightdata_headers']['font']

                # Add sheet title
                cell = worksheet.cell(row=2, column=1)
                cell.value = SHEETDATA[sheetname]['title']
                cell.font = SHEETSTYLES['sheet_titles']['font']

                # Add dropdown menus

                # Create data validation objects
                dvCamera = DataValidation(
                    type="list", formula1=f'"{",".join(DROPDOWNS["CAMERA"])}"'
                )
                dvLens = DataValidation(
                    type="list", formula1=f'"{",".join(DROPDOWNS["LENS"])}"'
                )
                dvFilter = DataValidation(
                    type="list", formula1=f'"{",".join(DROPDOWNS["FILTER"])}"'
                )
                dvInstrument = DataValidation(
                    type="list", formula1=f'"{",".join(DROPDOWNS["INSTRUMENT"])}"'
                )
                dvObservers = DataValidation(
                    type="list", formula1=f'"{",".join(DROPDOWNS["OBSERVERS"])}"'
                )

                # Add data validation objects to sheets
                worksheet.add_data_validation(dvCamera)
                worksheet.add_data_validation(dvLens)
                worksheet.add_data_validation(dvFilter)
                worksheet.add_data_validation(dvInstrument)
                worksheet.add_data_validation(dvObservers)

                # Set cell locations for each dropdown menu
                dvCamera.add(worksheet['M5'])
                dvLens.add(worksheet['O5'])
                dvFilter.add(worksheet['P5'])
                dvInstrument.add(worksheet['Q5'])
                dvObservers.add(worksheet['AA5'])
                dvObservers.add(worksheet['AB5'])
                dvObservers.add(worksheet['AC5'])
                dvObservers.add(worksheet['AD5'])

            else:
                # Add sheet title
                cell = worksheet.cell(row=1, column=1)
                cell.value = SHEETDATA[sheetname]['title']
                cell.font = SHEETSTYLES['sheet_titles']['font']
            

            # Add column headers and set column widths
            for i,colname in enumerate(SHEETDATA[sheetname]['colNames'],1):

                # Get desired column width
                colwidth = SHEETDATA[sheetname]['colWidths'][i-1]

                # Set cell value and style
                cell = worksheet.cell(row=4, column=i)
                cell.value = colname
                cell.font = SHEETSTYLES['column_headers']['font']
                cell.border = SHEETSTYLES['column_headers']['border']
                if colname in ['NARRATIVE','PLACE','COLLECTION_NOTES','CALIB_NOTES','NATSKY_FIT_NOTES']:
                    cell.alignment = SHEETSTYLES['column_headers']['alignment_lb']
                else:
                    cell.alignment = SHEETSTYLES['column_headers']['alignment_cb']
                colLetter = get_column_letter(i)
                worksheet.column_dimensions[colLetter].width = colwidth


def get_site_info(imageFile):
    '''
    Function that computes the Local Mean Time (LMT)
    and returns the site location name and observer
    names along with a formatted string containing 
    date and time info.
    '''

    # Load image header
    H = fits.getheader(imageFile)

    # Get site and observer names
    siteName = H['LOCATION']
    observers = H['OBSERVER']
    longitude = H['LONGITUD']
    latitude = H['LATITUDE']
    elevation = H['ELEVATIO']
    utc = H['DATE-OBS']
    tempc = (H['AMTEMP_F'] - 32) * 5./9

    # Get instrument details from FITS header
    fitsTelescope = H['TELESCOP']
    fitsInstrument = H['INSTRUME']
    fitsCamera = fitsInstrument.split(",")[0].strip()
    fitsLens = fitsInstrument.split(",")[1].strip()
    fitsFilter = H['FILTER'].strip()

    # Map FITS values to standard NIGHT METADATA entries
    if fitsTelescope in METADATA_MAPPINGS['INSTRUMENT'].keys():
        instrument = METADATA_MAPPINGS['INSTRUMENT'][fitsTelescope]
    else: instrument = None
    if fitsCamera in METADATA_MAPPINGS['CAMERA'].keys():
        camera = METADATA_MAPPINGS['CAMERA'][fitsCamera]
    else: camera = None
    if fitsLens in METADATA_MAPPINGS['LENS'].keys():
        lens = METADATA_MAPPINGS['LENS'][fitsLens]
    else: lens = None
    if camera is not None:
        if camera in METADATA_MAPPINGS['FILTER'][fitsFilter].keys():
            filterBatch = METADATA_MAPPINGS['FILTER'][fitsFilter][camera]
        else: filterBatch = None
    else: filterBatch = None

    # Convert observers string to list of observers and pad with empty strings
    observers = [x.strip() for x in observers.split(",")]
    while len(observers) < 4:
        observers.append("")

    # Get UTC date and time
    t = Time(utc, format='isot', scale='utc')
    utcDT = t.datetime
    utcDate = utcDT.strftime("%d-%b-%Y").lstrip("0")
    utcTime = utc.split("T")[1].lstrip("0")

    # Convert UTC to LMT using site longitude
    dayShift = 0
    hourfracUTC = t.ymdhms.hour + t.ymdhms.minute/60 + t.ymdhms.second/3600
    hourfracLMT = hourfracUTC + longitude/15
    if hourfracLMT < 0:
        hourfracLMT += 24
        dayShift -= 1

    # Approximate LMT midpoint of dataset
    midpointLMT = hourfracLMT + 0.16
    if midpointLMT > 24:
        midpointLMT -= 24
        dayShift += 1

    # Apply day shift
    t = t + dayShift*u.day
    lmtDT = t.datetime
    lmtDate = lmtDT.strftime("%d-%b-%Y").lstrip("0")

    # Create site info dict
    siteInfo = {
        'siteName': siteName,
        'observers': observers,
        'utcStartDate': utcDate,
        'utcStartTime': utcTime,
        'utcDT': utcDT,
        'longitude': longitude,
        'latitude': latitude,
        'elevation': elevation,
        'lmtMidDate': lmtDate,
        'lmtMidTime': midpointLMT,
        'lmtDT': lmtDT,
        'humidity': H['HUMID'],
        'windspeed': H['WIND_MPH'],
        'tempcelsius': tempc,
        'instrument': instrument,
        'camera': camera,
        'lens': lens,
        'filterName': fitsFilter,
        'filterBatch': filterBatch,
        'exptime': H['EXPTIME'],
        'ccdtemp': H['CCD-TEMP']
    }

    return siteInfo


def append_night_metadata(excelFile, siteInfo):

    # Sheet name
    sheetName = "NIGHT METADATA"

    # Add to NIGHT METADATA sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Load in zeropoint and platescale values
        calsetp = f"{filepath.calibdata}{siteInfo['datanight']}/"
        extinctionfile = f"{calsetp}extinction_fit_V.xlsx"
        extinctionData  = pd.read_excel(extinctionfile)
        zeropoint = extinctionData['zeropoint_default'].iloc[0]
        platescale = extinctionData['avg_scale'].mean()

        # Calculate image scale offset
        scaleOffset = 2.5*n.log10((platescale*60)**2)

        # Add processor name and processing date
        tnow = datetime.now()
        procDatetime = tnow.strftime("%m/%d/%y %H:%M").lstrip("0")
        pcell =  worksheet.cell(row=2, column=7)
        tcell =  worksheet.cell(row=2, column=8)
        pcell.value = siteInfo['processor']
        tcell.value = procDatetime
        pcell.font = SHEETSTYLES['nightdata_entries']['font']
        tcell.font = SHEETSTYLES['nightdata_entries']['font']
        pcell.alignment = SHEETSTYLES['nightdata_entries']['alignment_lb']
        tcell.alignment = SHEETSTYLES['nightdata_entries']['alignment_rb']

        # Set cell data values
        worksheet.cell(row=5, column=1 , value=siteInfo['datanight'])     # Data night
        worksheet.cell(row=5, column=2 , value=siteInfo['unitname'])      # NPS Unit Name
        worksheet.cell(row=5, column=3 , value=siteInfo['unitcode'])      # NPS Unit Code
        worksheet.cell(row=5, column=4 , value=siteInfo['longitude'])     # Longitude
        worksheet.cell(row=5, column=5 , value=siteInfo['latitude'])      # Latitude
        worksheet.cell(row=5, column=6 , value=siteInfo['elevation'])     # Elevation
        worksheet.cell(row=5, column=7 , value=siteInfo['siteName'])      # Site Name
        worksheet.cell(row=5, column=8 , value=siteInfo['utcDT'].date())  # UTC Start Date
        worksheet.cell(row=5, column=9 , value=siteInfo['utcDT'].time())  # UTC Start Time
        worksheet.cell(row=5, column=10, value=siteInfo['tempcelsius'])   # Temperature (C)
        worksheet.cell(row=5, column=11, value=siteInfo['humidity'])      # Relative Humidity (%)
        worksheet.cell(row=5, column=12, value=siteInfo['windspeed'])     # Wind speed (mph)
        worksheet.cell(row=5, column=14, value=siteInfo['ccdtemp'])       # Camera Temp
        worksheet.cell(row=5, column=18, value=45)                        # Number of Images
        worksheet.cell(row=5, column=19, value=siteInfo['exptime'])       # Exposure Time
        worksheet.cell(row=5, column=20, value=zeropoint)                 # Zeropoint
        worksheet.cell(row=5, column=21, value=scaleOffset)               # Image Scale Offset
        worksheet.cell(row=5, column=22, value=siteInfo['numsets'])       # Number of Sets
        # worksheet.cell(row=5, column=23, value=siteInfo['zlm'])         # Zenith Limiting Mag
        # worksheet.cell(row=5, column=24, value=siteInfo['bortle'])      # Bortle
        worksheet.cell(row=5, column=25, value=siteInfo['siteAlbedo'])    # Albedo
        # worksheet.cell(row=5, column=26, value=siteInfo['sqm'])         # SQM
        worksheet.cell(row=5, column=27, value=siteInfo['observers'][0])  # Observer 1
        worksheet.cell(row=5, column=28, value=siteInfo['observers'][1])  # Observer 2
        worksheet.cell(row=5, column=29, value=siteInfo['observers'][2])  # Observer 3
        worksheet.cell(row=5, column=30, value=siteInfo['observers'][3])  # Observer 4
        # worksheet.cell(row=5, column=31, value=siteInfo['narrative'])   # Narrative
        worksheet.cell(row=5, column=32, value=siteInfo['centralAZ'])     # Central Azimuth

        # Add instrument details when available
        if siteInfo['camera'] is not None:
            worksheet.cell(row=5, column=13, value=siteInfo['camera'])      # Camera name
        if siteInfo['lens'] is not None:
            worksheet.cell(row=5, column=15, value=siteInfo['lens'])        # Camera Lens name
        if siteInfo['filterBatch'] is not None:
            worksheet.cell(row=5, column=16, value=siteInfo['filterBatch']) # Filter batch ID
        if siteInfo['instrument'] is not None:
            worksheet.cell(row=5, column=17, value=siteInfo['instrument'])  # Telescope mount name

        # Set some cell number formats
        datestyle = SHEETSTYLES['datetime']['datestyle']
        timestyle = SHEETSTYLES['datetime']['timestyle']
        worksheet.cell(row=5, column=8).style = datestyle         # UTC Start Date
        worksheet.cell(row=5, column=9).style = timestyle         # UTC Start Time
        worksheet.cell(row=5, column=10).number_format = '0.0'    # Air temperature
        worksheet.cell(row=5, column=11).number_format = '0.0'    # Relative humidity
        worksheet.cell(row=5, column=20).number_format = '0.000'  # Zeropoint
        worksheet.cell(row=5, column=21).number_format = '0.000'  # Image Scale Offset
        worksheet.cell(row=5, column=25).number_format = '0.000'  # Albedo

        # Set cell styles
        ncol = len(SHEETDATA[sheetName]['colNames'])
        for i in range(ncol):
            cell = worksheet.cell(row=5, column=i+1)
            if SHEETDATA[sheetName]['colNames'][i] == 'NARRATIVE':
                cell.font = SHEETSTYLES['narrative']['font']
                cell.alignment = SHEETSTYLES['narrative']['alignment']
            else:
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_cities(excelFile, dnight):
    '''
    Function to append top 100 cities ranked by
    Walker Law values to CITIES sheet
    '''

    # Sheet name
    sheetName = "CITIES"

    # Load in cities excel sheet
    citiesFile = f"{filepath.calibdata}{dnight}/cities.xlsx"
    cities = pd.read_excel(citiesFile)

    # Add to CITIES sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Iterate over each row
        for i,row in cities.iterrows():

            # Break after 100 rows
            if i == 100:
                break
            
            # Set cell data values
            worksheet.cell(row=i+5, column=1, value=dnight)                       # Data night
            worksheet.cell(row=i+5, column=2, value=row['Place'])                 # City
            worksheet.cell(row=i+5, column=3, value=row['State'])                 # State
            worksheet.cell(row=i+5, column=4, value=row['Walkers Law'])           # Walkers Law
            worksheet.cell(row=i+5, column=5, value=row['Population'])            # Population
            worksheet.cell(row=i+5, column=6, value=row['Distance'])              # Distance (km)
            worksheet.cell(row=i+5, column=7, value=row['Bearing'])               # Bearing (deg)
            worksheet.cell(row=i+5, column=8, value=row['Half Width (Degrees)'])  # Half-width (deg)

            # Set some cell number formats
            worksheet.cell(row=i+5, column=4).number_format = '0.0000' # Walkers Law
            worksheet.cell(row=i+5, column=5).number_format = '#,###'  # Population
            worksheet.cell(row=i+5, column=6).number_format = '####'   # Distance
            worksheet.cell(row=i+5, column=7).number_format = '0.0'    # Bearing
            worksheet.cell(row=i+5, column=8).number_format = '0.0'    # Half-width (deg)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=i+5, column=j+1)
                if SHEETDATA[sheetName]['colNames'][j] == 'PLACE':
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_lb']
                else:
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_set_metadata(excelFile, dnight, sets):

    # Sheet name
    sheetName = "SET METADATA"

    # Add to SET METADATA sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])
            calsetp = f"{filepath.calibdata}{dnight}/S_{setnum:02d}/"

            # Get site data for first zenith image
            zenithImage = f"{calsetp}zenith1.fit"
            siteInfo = get_site_info(zenithImage)

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)                    # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)                    # Data set
            worksheet.cell(row=setnum+4, column=3 , value=siteInfo['utcDT'].date())  # UTC Start Date
            worksheet.cell(row=setnum+4, column=4 , value=siteInfo['utcDT'].time())  # UTC Start Time
            worksheet.cell(row=setnum+4, column=5 , value=siteInfo['lmtDT'].date())  # LMT Mid Date
            worksheet.cell(row=setnum+4, column=6 , value=siteInfo['lmtMidTime'])    # LMT Mid Time (hours)
            # worksheet.cell(row=setnum+4, column=7 , value=dnight)                  # Glare quality
            # worksheet.cell(row=setnum+4, column=8 , value=dnight)                  # Atmosphere quality
            # worksheet.cell(row=setnum+4, column=9 , value=dnight)                  # Collection quality
            # worksheet.cell(row=setnum+4, column=10, value=dnight)                  # Processing quality
            # worksheet.cell(row=setnum+4, column=11, value=dnight)                  # Reference Set
            # worksheet.cell(row=setnum+4, column=12, value=dnight)                  # Usable (Y/N)
            # worksheet.cell(row=setnum+4, column=13, value=dnight)                  # Clouds
            # worksheet.cell(row=setnum+4, column=14, value=dnight)                  # Plumes
            # worksheet.cell(row=setnum+4, column=15, value=dnight)                  # PCT20
            # worksheet.cell(row=setnum+4, column=16, value=dnight)                  # Collection Notes

            # Set some cell number/date formats
            datestyle = SHEETSTYLES['datetime']['datestyle']
            timestyle = SHEETSTYLES['datetime']['timestyle']
            worksheet.cell(row=setnum+4, column=3).style = datestyle       # UTC Start Date
            worksheet.cell(row=setnum+4, column=4).style = timestyle       # UTC Start Time
            worksheet.cell(row=setnum+4, column=5).style = datestyle       # LMT Start Date
            worksheet.cell(row=setnum+4, column=6).number_format = '0.00'  # LMT Mid Time

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                if SHEETDATA[sheetName]['colNames'][j] == 'COLLECTION_NOTES':
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_lb']
                else:
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_calibration(excelFile, dnight, sets):

    # Sheet name
    sheetName = "CALIBRATION"

    # Read in processing dataset list and skip rows where Process = No
    filelist = pd.read_excel(f"{filepath.processlist}filelist.xlsx")
    filelist = filelist.loc[filelist['Process'] == 'Yes'].reset_index(drop=True)

    # Get metadata for each dataset to be processed
    Dataset = filelist['Dataset'].values
    FlatV = filelist['Flat_V'].values
    Curve = filelist['Curve'].values

    # Set the flat and curve filenames and paths
    flatFile = FlatV[Dataset == dnight][0]
    flatPath = f"{filepath.flats}{flatFile}"
    curveName = Curve[Dataset == dnight][0]

    # Add to CALIBRATION sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])
            calsetp = f"{filepath.calibdata}{dnight}/"

            # Load bias drift data
            biasDriftFile = f"{calsetp}biasdrift_{setnum}.txt"
            biasDriftData = n.loadtxt(biasDriftFile)
            biasDrift = max(biasDriftData) - min(biasDriftData)

            # Load pointing error data
            pterrFile = f"{calsetp}pointerr_{setnum}.txt"
            pterrData = n.loadtxt(pterrFile)
            pterrMax = max(pterrData[:,7])
            pterrAvg = n.mean(pterrData[:,7])

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1, value=dnight)     # Data night
            worksheet.cell(row=setnum+4, column=2, value=setnum)     # Data set
            worksheet.cell(row=setnum+4, column=3, value=flatPath)   # Flat File
            worksheet.cell(row=setnum+4, column=4, value=curveName)  # Linearity Curve
            worksheet.cell(row=setnum+4, column=5, value=biasDrift)  # Bias Drift
            worksheet.cell(row=setnum+4, column=6, value=pterrMax)   # Max Pointing Err
            worksheet.cell(row=setnum+4, column=7, value=pterrAvg)   # Avg Pointing Err

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=5).number_format = '0.0'   # Bias Drift
            worksheet.cell(row=setnum+4, column=6).number_format = '0.00'  # Max Pointing Err
            worksheet.cell(row=setnum+4, column=7).number_format = '0.00'  # Avg Pointing Err

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                if SHEETDATA[sheetName]['colNames'][j] == 'CALIB_NOTES':
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_lb']
                else:
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_extinction(excelFile, dnight, sets):

    # Sheet name
    sheetName = "EXTINCTION"

    # Load in extinction data
    extinctionFile = f"{filepath.calibdata}{dnight}/extinction_fit_V.xlsx"
    extinctionData  = pd.read_excel(extinctionFile)
    imgSolved = extinctionData['img_solved'].values
    starsFit = extinctionData['num_star_used'].values
    starsRej = extinctionData['num_star_rejected'].values
    zpFree = extinctionData['zeropoint_free'].values
    zpFreeErr = extinctionData['zeropoint_free_err'].values
    zpFixed = extinctionData['zeropoint_default'].values
    extFixed = extinctionData['extinction_fixedZ'].values
    colorFixed = extinctionData['color_coeff_default'].values
    colorFree = extinctionData['color_coeff_free'].values

    # Add to EXTINCTION sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)                # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)                # Data set
            worksheet.cell(row=setnum+4, column=3 , value=imgSolved[setnum-1])   # Images solved
            worksheet.cell(row=setnum+4, column=4 , value=starsFit[setnum-1])    # Stars used
            worksheet.cell(row=setnum+4, column=5 , value=starsRej[setnum-1])    # Stars rejected
            worksheet.cell(row=setnum+4, column=6 , value=zpFreeErr[setnum-1])   # Bestfit Zeropoint Error
            worksheet.cell(row=setnum+4, column=7 , value=zpFree[setnum-1])      # Bestfit Zeropoint
            worksheet.cell(row=setnum+4, column=8 , value=-colorFree[setnum-1])  # Bestfit Color Coeff
            worksheet.cell(row=setnum+4, column=9 , value=zpFixed[setnum-1])     # Default Zeropoint
            worksheet.cell(row=setnum+4, column=10, value=colorFixed[setnum-1])  # Default Color Coeff
            worksheet.cell(row=setnum+4, column=11, value=-extFixed[setnum-1])   # Bestfit Ext Coeff (Fixed ZP)

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.000'  # Bestfit Zeropoint Error
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.000'  # Bestfit Zeropoint
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.000'  # Bestfit Color Coeff
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.000'  # Default Zeropoint
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'  # Default Color Coeff
            worksheet.cell(row=setnum+4, column=11).number_format = '0.00'   # Bestfit Ext Coeff (Fixed ZP)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_coordinates(excelFile, dnight, sets):

    # Sheet name
    sheetName = "IMG COORDS"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])
            calsetp = f"{filepath.calibdata}{dnight}/"

            # Loadin in pointing error file for Alt/Az coords
            pterrFile = f"{calsetp}pointerr_{setnum}.txt"
            pterr = n.loadtxt(pterrFile)

            # Load in image coordinates file
            coordFile = f"{calsetp}coordinates_{setnum}.txt"
            coord = n.loadtxt(coordFile)
            numImages = len(coord)

            # Iterate over each image
            for i in range(numImages):
                
                # Set image number
                imgnum = i+1

                # Set cell data values
                row = (setnum-1) * numImages + imgnum + 4
                worksheet.cell(row=row, column=1 , value=dnight)           # Data night
                worksheet.cell(row=row, column=2 , value=setnum)           # Data set
                worksheet.cell(row=row, column=3 , value=imgnum)           # Image number
                worksheet.cell(row=row, column=4 , value=pterr[i,3])       # Azimuth (deg)
                worksheet.cell(row=row, column=5 , value=pterr[i,4])       # Altitude (deg)
                worksheet.cell(row=row, column=6 , value=coord[i,7]/15)    # RA (hr)
                worksheet.cell(row=row, column=7 , value=coord[i,8])       # Dec (deg)
                worksheet.cell(row=row, column=8 , value=coord[i,2])       # Gal Longitude (deg)
                worksheet.cell(row=row, column=9 , value=coord[i,3])       # Gal Latitude (deg)
                worksheet.cell(row=row, column=10, value=coord[i,5])       # Ecl Longitude (deg)
                worksheet.cell(row=row, column=11, value=coord[i,6])       # Ecl Latitude (deg)

                # Set some cell number/date formats
                worksheet.cell(row=row, column=4 ).number_format = '0.00'  # Azimuth (deg)
                worksheet.cell(row=row, column=5 ).number_format = '0.00'  # Altitude (deg)
                worksheet.cell(row=row, column=6 ).number_format = '0.00'  # RA (hr)
                worksheet.cell(row=row, column=7 ).number_format = '0.00'  # Dec (deg)
                worksheet.cell(row=row, column=8 ).number_format = '0.00'  # Gal Longitude (deg)
                worksheet.cell(row=row, column=9 ).number_format = '0.00'  # Gal Latitude (deg)
                worksheet.cell(row=row, column=10).number_format = '0.00'  # Ecl Longitude (deg)
                worksheet.cell(row=row, column=11).number_format = '0.00'  # Ecl Latitude (deg)

                # Set cell styles
                ncol = len(SHEETDATA[sheetName]['colNames'])
                for j in range(ncol):
                    cell = worksheet.cell(row=row, column=j+1)
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_natsky_params(excelFile, dnight, sets):

    # Sheet name
    sheetName = "NATSKY"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])
            calsetp = f"{filepath.calibdata}{dnight}/"

            # Load in natural sky model parameters
            natskyFile = f"{calsetp}natsky_model_params.xlsx"
            natskyParams = pd.read_excel(
                natskyFile, sheet_name="Model_Parameters"
            )

            # Extract individual model parameters
            setIdx = (natskyParams['Data Set'] == setnum)
            airglowHeight = natskyParams[setIdx]['Emitting Layer Height (km)'].iloc[0]
            siteElevation = natskyParams[setIdx]['Site Elevation (m)'].iloc[0]
            extCoeff = natskyParams[setIdx]['Extinction Coefficient'].iloc[0]
            airglowZenithNl = natskyParams[setIdx]['Zenith Airglow (nL)'].iloc[0]
            airglowExt = natskyParams[setIdx]['Airglow Extinction Constant'].iloc[0]
            adlFactor = natskyParams[setIdx]['A.D.L. Multiplier'].iloc[0]
            zodiacalExt = natskyParams[setIdx]['Zodiacal Extinction Constant'].iloc[0]
            galacticExt = natskyParams[setIdx]['Galactic Extinction Constant'].iloc[0]
            fitQuality = natskyParams[setIdx]['Quality Flag (0-5)'].iloc[0]
            natskyNotes = natskyParams[setIdx]['Notes'].iloc[0]

            # Unit conversions
            airglowZenithMccd = nl_to_mccd(airglowZenithNl)  # nL to uCd

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)             # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)             # Data set
            worksheet.cell(row=setnum+4, column=3 , value=airglowHeight)      # Flat File
            worksheet.cell(row=setnum+4, column=4 , value=siteElevation)      # Linearity Curve
            worksheet.cell(row=setnum+4, column=5 , value=extCoeff)           # Extinction Coeff
            worksheet.cell(row=setnum+4, column=6 , value=airglowZenithMccd)  # Max Pointing Err
            worksheet.cell(row=setnum+4, column=7 , value=airglowExt)         # Airglow extinction constant
            worksheet.cell(row=setnum+4, column=8 , value=adlFactor)          # A.D.L. multiplier
            worksheet.cell(row=setnum+4, column=9 , value=zodiacalExt)        # Zodiacal extinction constant
            worksheet.cell(row=setnum+4, column=10, value=galacticExt)        # Galactic extinction constant
            worksheet.cell(row=setnum+4, column=11, value=fitQuality)         # Fit quality rating (0-5)
            worksheet.cell(row=setnum+4, column=12, value=natskyNotes)        # Notes

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.000'   # Extinction Coeff
            worksheet.cell(row=setnum+4, column=6 ).number_format = '#,###'   # Max Pointing Err
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.0'     # Airglow extinction constant
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.0'     # A.D.L. multiplier
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.0'     # Zodiacal extinction constant
            worksheet.cell(row=setnum+4, column=10).number_format = '0.0'     # Galactic extinction constant

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                if SHEETDATA[sheetName]['colNames'][j] == 'NATSKY_FIT_NOTES':
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_lb']
                else:
                    cell.font = SHEETSTYLES['data_fields']['font']
                    cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_photometryV2(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "V2 PHOTOMETRY"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed sky quality metrics
            sqMetrics = metrics['skyquality']
            sqIndex = (
                (sqMetrics['dataset'] == setnum) &
                (sqMetrics['filter'] == 'V')
            )
            zenithMag = sqMetrics[sqIndex]['zenith_mag'].iloc[0]
            allskyMag = sqMetrics[sqIndex]['allsky_mag'].iloc[0]
            za70Mag = sqMetrics[sqIndex]['za70_mag'].iloc[0]
            brightestMag = sqMetrics[sqIndex]['brightest_mag'].iloc[0]
            faintestMag = sqMetrics[sqIndex]['faintest_mag'].iloc[0]
            sqm = sqMetrics[sqIndex]['SQM_synthetic'].iloc[0]
            scalarIllum = sqMetrics[sqIndex]['scalar_illum'].iloc[0]

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1, value=dnight)         # Data night
            worksheet.cell(row=setnum+4, column=2, value=setnum)         # Data set
            worksheet.cell(row=setnum+4, column=3, value=zenithMag)      # Zenith Mag
            worksheet.cell(row=setnum+4, column=4, value=allskyMag)      # Allsky Total Mag
            worksheet.cell(row=setnum+4, column=5, value=za70Mag)        # ZA70 Total Mag
            worksheet.cell(row=setnum+4, column=6, value=brightestMag)   # Brightest Mag
            worksheet.cell(row=setnum+4, column=7, value=faintestMag)    # Faintest Mag
            worksheet.cell(row=setnum+4, column=8, value=sqm)            # Synthetic SQM
            worksheet.cell(row=setnum+4, column=9, value=scalarIllum)    # Scalar illuminance

            # # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3).number_format = '0.00'   # Zenith Mag
            worksheet.cell(row=setnum+4, column=4).number_format = '0.00'   # Allsky Total Mag
            worksheet.cell(row=setnum+4, column=5).number_format = '0.00'   # ZA70 Total Mag
            worksheet.cell(row=setnum+4, column=6).number_format = '0.00'   # Brightest Mag
            worksheet.cell(row=setnum+4, column=7).number_format = '0.00'   # Faintest Mag
            worksheet.cell(row=setnum+4, column=8).number_format = '0.00'   # Synthetic SQM
            worksheet.cell(row=setnum+4, column=9).number_format = '0.000'  # Scalar illuminance

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_photometryV4(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "V4 PHOTOMETRY"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            iaMetrics = metrics['illumall']
            sqMetrics = metrics['skyquality']
            svMetrics = metrics['starsvis']
            iaIndex = (
                (iaMetrics['dataset'] == setnum) &
                (iaMetrics['filter'] == 'V')
            )
            sqIndex = (
                (sqMetrics['dataset'] == setnum) &
                (sqMetrics['filter'] == 'V')
            )
            svIndex = (
                (svMetrics['dataset'] == setnum) &
                (svMetrics['filter'] == 'V')
            )
            meanLumAllskyNl = iaMetrics[iaIndex]['skyave1'].iloc[0]
            zenithLumMag = sqMetrics[sqIndex]['zenith_mag'].iloc[0]
            brightestLumNl = iaMetrics[iaIndex]['skymax0'].iloc[0]
            totalIllumMlux = iaMetrics[iaIndex]['totalill0'].iloc[0]
            horizontalIllumMlux = iaMetrics[iaIndex]['horizs'].iloc[0]
            numStarsFlat = svMetrics[svIndex]['Nstar_flat_horizon'].iloc[0]
            numStarsOnSky = svMetrics[svIndex]['Nstar_obs_horizon'].iloc[0]
            numStarsExtincted = svMetrics[svIndex]['Nstar_vis_noBkg'].iloc[0]
            numStarsVisNatsky = svMetrics[svIndex]['Nstar_vis_natsky'].iloc[0]
            numStarsVisPolluted = svMetrics[svIndex]['Nstar_vis_polluted'].iloc[0]
            fracStarsVis = 100 * numStarsVisPolluted / numStarsVisNatsky
            scalarIllum = iaMetrics[iaIndex]['skyscalar'].iloc[0] * metrics['albedo'] / 4

            # Find maximum vertical illuminance
            vertColumns = [col for col in iaMetrics.columns if 'vert-' in col]
            maxVertIllumMlux = max(iaMetrics[iaIndex][vertColumns].values[0])

            # Unit conversions
            meanLumAllskyMccd = nl_to_mccd(meanLumAllskyNl)  # nL to uCd
            meanLumAllskyMag =  nl_to_mags(meanLumAllskyNl)  # nL to mag/arcsec^2
            zenithLumMccd = mags_to_mccd(zenithLumMag)       # mag/arcsec^2 to uCd
            brightestLumMccd = nl_to_mccd(brightestLumNl)    # nL to uCd
            brightestLumMag = nl_to_mags(brightestLumNl)     # nL to mag/arcsec^2
            totalIllumMag = mlux_to_mag(totalIllumMlux)      # mlux to mag

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)               # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)               # Data set
            worksheet.cell(row=setnum+4, column=3 , value=meanLumAllskyMag)     # Mean Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=4 , value=meanLumAllskyMccd)    # Mean Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 , value=zenithLumMag)         # Zenith Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=6 , value=zenithLumMccd)        # Zenith Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 , value=brightestLumMag)      # Brightest Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=8 , value=brightestLumMccd)     # Brightest Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 , value=totalIllumMag)        # All-sky luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=10, value=totalIllumMlux)       # All-sky luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=11, value=horizontalIllumMlux)  # Horizontal Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12, value=maxVertIllumMlux)     # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=13, value=numStarsFlat)         # Num Stars to Flat Horizon
            worksheet.cell(row=setnum+4, column=14, value=numStarsOnSky)        # Num Stars to Observed Horizon
            worksheet.cell(row=setnum+4, column=15, value=numStarsExtincted)    # Num Stars after Extinction
            worksheet.cell(row=setnum+4, column=16, value=numStarsVisNatsky)    # Num Stars Visible, Natural Sky
            worksheet.cell(row=setnum+4, column=17, value=numStarsVisPolluted)  # Num Stars Visible, Polluted Sky
            worksheet.cell(row=setnum+4, column=18, value=fracStarsVis)         # Fraction of Stars Visible
            worksheet.cell(row=setnum+4, column=19, value=scalarIllum)          # Scalar Illuminance

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.00'    # Mean Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.0'     # Mean Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'    # Zenith Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.0'     # Zenith Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.00'    # Brightest Luminance (mag/arcesc^2)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.0'     # Brightest Luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.00'    # All-sky luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'   # All-sky luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=11).number_format = '0.000'   # Horizontal Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.000'   # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=13).number_format = '####'    # Num Stars to Flat Horizon
            worksheet.cell(row=setnum+4, column=14).number_format = '####'    # Num Stars to Observed Horizon
            worksheet.cell(row=setnum+4, column=15).number_format = '####'    # Num Stars after Extinction
            worksheet.cell(row=setnum+4, column=16).number_format = '####'    # Num Stars Visible, Natural Sky
            worksheet.cell(row=setnum+4, column=17).number_format = '####'    # Num Stars Visible, Polluted Sky
            worksheet.cell(row=setnum+4, column=18).number_format = '0.00'    # Fraction of Stars Visible
            worksheet.cell(row=setnum+4, column=19).number_format = '0.0000'  # Scalar Illuminance

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_lp_allsky(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "LP"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            sgMetrics = metrics['skyglow']
            sqMetrics = metrics['skyquality']
            nsMetrics = metrics['natsky_art']
            sgIndex = (
                (sgMetrics['dataset'] == setnum) &
                (sgMetrics['filter'] == 'V')
            )
            sqIndex = (
                (sqMetrics['dataset'] == setnum) &
                (sqMetrics['filter'] == 'V')
            )
            nsIndex = (
                (nsMetrics['Data Set'] == setnum)
            )
            siteALR = metrics['alr']
            sqi = sqMetrics[sqIndex]['SQI_allsky'].iloc[0]
            meanLumAllskyMlux = sgMetrics[sgIndex]['hemis0'].iloc[0]
            totalLumAllskyMlux = sgMetrics[sgIndex]['totalill0'].iloc[0]
            horizIllumAllskyMlux = sgMetrics[sgIndex]['horizs0'].iloc[0]
            brightestLumNl = sgMetrics[sgIndex]['skymax0'].iloc[0]
            zenithLumNl = nsMetrics[nsIndex]['Zenith'].iloc[0]

            # Get vertical illuminance stats
            vertColumnsAll = [col for col in sgMetrics.columns if 'vert-' in col]
            vertColumns = [col for col in vertColumnsAll if col[-1] == '0']
            maxVertIllumMlux = max(sgMetrics[sgIndex][vertColumns].values[0])
            minVertIllumMlux = min(sgMetrics[sgIndex][vertColumns].values[0])
            meanVertIllumMlux = n.mean(sgMetrics[sgIndex][vertColumns].values[0])

            # Unit conversions
            meanLumAllskyNl = mlux_to_nl(meanLumAllskyMlux)       # mlux to nL
            meanLumAllskyMccd = nl_to_mccd(meanLumAllskyNl)       # nL to uCd
            lprAllsky = meanLumAllskyNl / 78                      # All-sky LPR (78 nL = natural condition)
            totalLumAllskyMag = mlux_to_mag(totalLumAllskyMlux)   # mlux to mag
            lprVertMax = maxVertIllumMlux / 0.4                   # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMin = minVertIllumMlux / 0.4                   # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMean = meanVertIllumMlux / 0.4                 # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprHorizontal = horizIllumAllskyMlux / 0.8            # Horizontal Illum LPR (0.8 mlux = natural condition)
            brightestLumMccd = nl_to_mccd(brightestLumNl)         # nL to uCd
            lprBrightest = brightestLumNl / 54                    # Brightest LPR (54 nL = natural condition)
            zenithLumMccd = nl_to_mccd(zenithLumNl)               # nL to uCd
            lprZenith = zenithLumNl / 54                          # Zenith LPR (54 nL = natural condition)

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)               # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)               # Data set
            worksheet.cell(row=setnum+4, column=3 , value=sqi)                  # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 , value=lprAllsky)            # All-sky Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 , value=totalLumAllskyMag)    # All-sky luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 , value=totalLumAllskyMlux)   # All-sky luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 , value=maxVertIllumMlux)     # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 , value=lprVertMax)           # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 , value=meanVertIllumMlux)    # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10, value=lprVertMean)          # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11, value=minVertIllumMlux)     # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12, value=lprVertMin)           # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13, value=horizIllumAllskyMlux) # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14, value=lprHorizontal)        # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15, value=brightestLumMccd)     # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16, value=lprBrightest)         # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17, value=zenithLumMccd)        # Zenith luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=18, value=lprZenith)            # Zenith luminance (LPR)
            worksheet.cell(row=setnum+4, column=19, value=meanLumAllskyMccd)    # Mean all-sky luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=20, value=siteALR)              # Modeled ALR for site

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.0'       # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.000'     # All-sky Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'      # All-sky luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.000'     # All-sky luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.000'     # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.000'     # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.000'     # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'     # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11).number_format = '0.000'     # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.000'     # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13).number_format = '0.000'     # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14).number_format = '0.000'     # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15).number_format = '####'      # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16).number_format = '0.00'      # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17).number_format = '####'      # Zenith luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=18).number_format = '0.000'     # Zenith luminance (LPR)
            worksheet.cell(row=setnum+4, column=19).number_format = '####'      # Mean all-sky luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=20).number_format = '0.000'     # Modeled ALR for site

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_lp_za80(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "LP80"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            sgMetrics = metrics['skyglow']
            sqMetrics = metrics['skyquality']
            sgIndex = (
                (sgMetrics['dataset'] == setnum) &
                (sgMetrics['filter'] == 'V')
            )
            sqIndex = (
                (sqMetrics['dataset'] == setnum) &
                (sqMetrics['filter'] == 'V')
            )
            sqi = sqMetrics[sqIndex]['SQI_ZA80'].iloc[0]
            meanLumZA80Mlux = sgMetrics[sgIndex]['hemis1'].iloc[0]
            totalLumZA80Mlux = sgMetrics[sgIndex]['totalill1'].iloc[0]
            horizIllumZA80Mlux = sgMetrics[sgIndex]['horizs1'].iloc[0]
            brightestLumNl = sgMetrics[sgIndex]['skymax1'].iloc[0]

            # Get vertical illuminance stats
            vertColumnsAll = [col for col in sgMetrics.columns if 'vert-' in col]
            vertColumns = [col for col in vertColumnsAll if col[-1] == '1']
            maxVertIllumMlux = max(sgMetrics[sgIndex][vertColumns].values[0])
            minVertIllumMlux = min(sgMetrics[sgIndex][vertColumns].values[0])
            meanVertIllumMlux = n.mean(sgMetrics[sgIndex][vertColumns].values[0])

            # Unit conversions
            meanLumZA80Nl = mlux_to_nl(meanLumZA80Mlux)       # mlux to nL
            meanLumZA80Mccd = nl_to_mccd(meanLumZA80Nl)       # nL to uCd
            lprZA80 = meanLumZA80Nl / 78                      # ZA-80 LPR (78 nL = natural condition)
            totalLumZA80Mag = mlux_to_mag(totalLumZA80Mlux)   # mlux to mag
            lprVertMax = maxVertIllumMlux / 0.4               # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMin = minVertIllumMlux / 0.4               # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMean = meanVertIllumMlux / 0.4             # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprHorizontal = horizIllumZA80Mlux / 0.8          # Horizontal Illum LPR (0.8 mlux = natural condition)
            brightestLumMccd = nl_to_mccd(brightestLumNl)     # nL to uCd
            lprBrightest = brightestLumNl / 54                # Brightest LPR (54 nL = natural condition)

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)             # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)             # Data set
            worksheet.cell(row=setnum+4, column=3 , value=sqi)                # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 , value=lprZA80)            # ZA-80 Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 , value=totalLumZA80Mag)    # ZA-80 luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 , value=totalLumZA80Mlux)   # ZA-80 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 , value=maxVertIllumMlux)   # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 , value=lprVertMax)         # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 , value=meanVertIllumMlux)  # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10, value=lprVertMean)        # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11, value=minVertIllumMlux)   # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12, value=lprVertMin)         # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13, value=horizIllumZA80Mlux) # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14, value=lprHorizontal)      # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15, value=brightestLumMccd)   # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16, value=lprBrightest)       # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17, value=meanLumZA80Mccd)    # Mean ZA-80 luminance (micro-Candela)

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.0'       # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.000'     # ZA-80 Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'      # ZA-80 luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.000'     # ZA-80 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.000'     # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.000'     # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.000'     # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'     # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11).number_format = '0.000'     # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.000'     # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13).number_format = '0.000'     # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14).number_format = '0.000'     # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15).number_format = '####'      # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16).number_format = '0.00'      # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17).number_format = '####'      # Mean ZA-80 luminance (micro-Candela)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_lp_za70(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "LP70"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            sgMetrics = metrics['skyglow']
            sqMetrics = metrics['skyquality']
            sgIndex = (
                (sgMetrics['dataset'] == setnum) &
                (sgMetrics['filter'] == 'V')
            )
            sqIndex = (
                (sqMetrics['dataset'] == setnum) &
                (sqMetrics['filter'] == 'V')
            )
            sqi = sqMetrics[sqIndex]['SQI_ZA70'].iloc[0]
            meanLumZA70Mlux = sgMetrics[sgIndex]['hemis2'].iloc[0]
            totalLumZA70Mlux = sgMetrics[sgIndex]['totalill2'].iloc[0]
            horizIllumZA70Mlux = sgMetrics[sgIndex]['horizs2'].iloc[0]
            brightestLumNl = sgMetrics[sgIndex]['skymax2'].iloc[0]

            # Get vertical illuminance stats
            vertColumnsAll = [col for col in sgMetrics.columns if 'vert-' in col]
            vertColumns = [col for col in vertColumnsAll if col[-1] == '2']
            maxVertIllumMlux = max(sgMetrics[sgIndex][vertColumns].values[0])
            minVertIllumMlux = min(sgMetrics[sgIndex][vertColumns].values[0])
            meanVertIllumMlux = n.mean(sgMetrics[sgIndex][vertColumns].values[0])

            # Unit conversions
            meanLumZA70Nl = mlux_to_nl(meanLumZA70Mlux)       # mlux to nL
            meanLumZA70Mccd = nl_to_mccd(meanLumZA70Nl)       # nL to uCd
            lprZA70 = meanLumZA70Nl / 78                      # ZA-70 LPR (78 nL = natural condition)
            totalLumZA70Mag = mlux_to_mag(totalLumZA70Mlux)   # mlux to mag
            lprVertMax = maxVertIllumMlux / 0.4               # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMin = minVertIllumMlux / 0.4               # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprVertMean = meanVertIllumMlux / 0.4             # Max Vertical Illum LPR (0.4 mlux = natural condition)
            lprHorizontal = horizIllumZA70Mlux / 0.8          # Horizontal Illum LPR (0.8 mlux = natural condition)
            brightestLumMccd = nl_to_mccd(brightestLumNl)     # nL to uCd
            lprBrightest = brightestLumNl / 54                # Brightest LPR (54 nL = natural condition)

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)             # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)             # Data set
            worksheet.cell(row=setnum+4, column=3 , value=sqi)                # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 , value=lprZA70)            # ZA-70 Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 , value=totalLumZA70Mag)    # ZA-70 luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 , value=totalLumZA70Mlux)   # ZA-70 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 , value=maxVertIllumMlux)   # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 , value=lprVertMax)         # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 , value=meanVertIllumMlux)  # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10, value=lprVertMean)        # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11, value=minVertIllumMlux)   # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12, value=lprVertMin)         # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13, value=horizIllumZA70Mlux) # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14, value=lprHorizontal)      # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15, value=brightestLumMccd)   # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16, value=lprBrightest)       # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17, value=meanLumZA70Mccd)    # Mean ZA-70 luminance (micro-Candela)

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.0'       # Sky Quality Index
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.000'     # ZA-80 Light pollution ratio (LPR)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'      # ZA-80 luminous emittance (mag)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.000'     # ZA-80 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.000'     # Max Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.000'     # Max Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.000'     # Mean Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'     # Mean Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=11).number_format = '0.000'     # Min Vertical Illum (milli-Lux)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.000'     # Min Vertical Illum (LPR))
            worksheet.cell(row=setnum+4, column=13).number_format = '0.000'     # Horizontal Illum (mlux)
            worksheet.cell(row=setnum+4, column=14).number_format = '0.000'     # Horizontal Illum (LPR)
            worksheet.cell(row=setnum+4, column=15).number_format = '####'      # Brightest luminance (micro-Candela)
            worksheet.cell(row=setnum+4, column=16).number_format = '0.00'      # Brightest luminance (LPR)
            worksheet.cell(row=setnum+4, column=17).number_format = '####'      # Mean ZA-80 luminance (micro-Candela)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_zones(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "ZONES"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            sgMetrics = metrics['skyglow']
            sgIndex = (
                (sgMetrics['dataset'] == setnum) &
                (sgMetrics['filter'] == 'V')
            )
            meanLumNlZone1 = sgMetrics[sgIndex]['zoneAve0'].iloc[0]
            meanLumNlZone2 = sgMetrics[sgIndex]['zoneAve1'].iloc[0]
            meanLumNlZone3 = sgMetrics[sgIndex]['zoneAve2'].iloc[0]
            meanLumNlZone4 = sgMetrics[sgIndex]['zoneAve3'].iloc[0]
            meanLumNlZone5 = sgMetrics[sgIndex]['zoneAve4'].iloc[0]
            countsZone1 = sgMetrics[sgIndex]['zoneMax0'].iloc[0]
            countsZone2 = sgMetrics[sgIndex]['zoneMax1'].iloc[0]
            countsZone3 = sgMetrics[sgIndex]['zoneMax2'].iloc[0]
            countsZone4 = sgMetrics[sgIndex]['zoneMax3'].iloc[0]
            countsZone5 = sgMetrics[sgIndex]['zoneMax4'].iloc[0]

            # Unit conversions
            lumEmitMluxZone1 = nl_to_mlux(meanLumNlZone1) * countsZone1   # nL to mlux
            lumEmitMluxZone2 = nl_to_mlux(meanLumNlZone2) * countsZone2   # nL to mlux
            lumEmitMluxZone3 = nl_to_mlux(meanLumNlZone3) * countsZone3   # nL to mlux
            lumEmitMluxZone4 = nl_to_mlux(meanLumNlZone4) * countsZone4   # nL to mlux
            lumEmitMluxZone5 = nl_to_mlux(meanLumNlZone5) * countsZone5   # nL to mlux

            # Calculate percentage of sky luminous emittance in each zone
            zoneSum = lumEmitMluxZone1 + lumEmitMluxZone2 + lumEmitMluxZone3 + lumEmitMluxZone4 + lumEmitMluxZone5
            pctZone1 = 100 * lumEmitMluxZone1 / zoneSum
            pctZone2 = 100 * lumEmitMluxZone2 / zoneSum
            pctZone3 = 100 * lumEmitMluxZone3 / zoneSum
            pctZone4 = 100 * lumEmitMluxZone4 / zoneSum
            pctZone5 = 100 * lumEmitMluxZone5 / zoneSum

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)            # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)            # Data set
            worksheet.cell(row=setnum+4, column=3 , value=meanLumNlZone1)    # Zone 1 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=4 , value=lumEmitMluxZone1)  # Zone 1 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=5 , value=pctZone1)          # Zone 1 Sky Percentage
            worksheet.cell(row=setnum+4, column=6 , value=meanLumNlZone2)    # Zone 2 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=7 , value=lumEmitMluxZone2)  # Zone 2 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 , value=pctZone2)          # Zone 2 Sky Percentage
            worksheet.cell(row=setnum+4, column=9 , value=meanLumNlZone3)    # Zone 3 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=10, value=lumEmitMluxZone3)  # Zone 3 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=11, value=pctZone3)          # Zone 3 Sky Percentage
            worksheet.cell(row=setnum+4, column=12, value=meanLumNlZone4)    # Zone 4 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=13, value=lumEmitMluxZone4)  # Zone 4 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=14, value=pctZone4)          # Zone 4 Sky Percentage
            worksheet.cell(row=setnum+4, column=15, value=meanLumNlZone5)    # Zone 5 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=16, value=lumEmitMluxZone5)  # Zone 5 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=17, value=pctZone5)          # Zone 5 Sky Percentage

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '####'   # Zone 1 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.000'  # Zone 1 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.0'    # Zone 1 Sky Percentage
            worksheet.cell(row=setnum+4, column=6 ).number_format = '####'   # Zone 2 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.000'  # Zone 2 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.0'    # Zone 2 Sky Percentage
            worksheet.cell(row=setnum+4, column=9 ).number_format = '####'   # Zone 3 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.000'  # Zone 3 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=11).number_format = '0.0'    # Zone 3 Sky Percentage
            worksheet.cell(row=setnum+4, column=12).number_format = '####'   # Zone 4 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=13).number_format = '0.000'  # Zone 4 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=14).number_format = '0.0'    # Zone 4 Sky Percentage
            worksheet.cell(row=setnum+4, column=15).number_format = '####'   # Zone 5 mean Luminance (nL)
            worksheet.cell(row=setnum+4, column=16).number_format = '0.000'  # Zone 5 luminous emittance (milli-Lux)
            worksheet.cell(row=setnum+4, column=17).number_format = '0.0'    # Zone 5 Sky Percentage

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_percentiles_all(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "V4 PERCENTILES ALL"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            nsMetrics = metrics['natsky_all']
            nsIndex = (
                (nsMetrics['Data Set'] == setnum)
            )
            pctDeg0p5 = nl_to_mccd(nsMetrics[nsIndex]['0.5 Degree Percentile'].iloc[0])
            pctDeg1p0 = nl_to_mccd(nsMetrics[nsIndex]['1.0 Degree Percentile'].iloc[0])
            pct99 = nl_to_mccd(nsMetrics[nsIndex]['99th Percentile'].iloc[0])
            pct98 = nl_to_mccd(nsMetrics[nsIndex]['98th Percentile'].iloc[0])
            pct95 = nl_to_mccd(nsMetrics[nsIndex]['95th Percentile'].iloc[0])
            pct90 = nl_to_mccd(nsMetrics[nsIndex]['90th Percentile'].iloc[0])
            pct80 = nl_to_mccd(nsMetrics[nsIndex]['80th Percentile'].iloc[0])
            pct70 = nl_to_mccd(nsMetrics[nsIndex]['70th Percentile'].iloc[0])
            pct60 = nl_to_mccd(nsMetrics[nsIndex]['60th Percentile'].iloc[0])
            pct50 = nl_to_mccd(nsMetrics[nsIndex]['Median'].iloc[0])
            pct1 = nl_to_mccd(nsMetrics[nsIndex]['1st Percentile'].iloc[0])
            pctMin = nl_to_mccd(nsMetrics[nsIndex]['Minimum (0.05 Percentile)'].iloc[0])

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)           # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)           # Data set
            worksheet.cell(row=setnum+4, column=3 , value=pctDeg0p5)        # 0.5 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=4 , value=pctDeg1p0)        # 1.0 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 , value=pct99)            # 99th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=6 , value=pct98)            # 98th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 , value=pct95)            # 95th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=8 , value=pct90)            # 90th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 , value=pct80)            # 80th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=10, value=pct70)            # 70th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=11, value=pct60)            # 60th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=12, value=pct50)            # 50th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=13, value=pct1)             #  1st Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=14, value=pctMin)           # 0.05 Percentile (micro-Candela)

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.00'  # 0.5 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.00'  # 1.0 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'  # 99th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.00'  # 98th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.00'  # 95th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.00'  # 90th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.00'  # 80th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.00'  # 70th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=11).number_format = '0.00'  # 60th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.00'  # 50th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=13).number_format = '0.00'  #  1st Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=14).number_format = '0.00'  # 0.05 Percentile (micro-Candela)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


def append_percentiles_lp(excelFile, dnight, sets, metrics):

    # Sheet name
    sheetName = "V4 PERCENTILES LP"

    # Add to IMG COORDS sheet
    with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:

        # Grab the relevant worksheet
        worksheet = writer.sheets[sheetName]

        # Loop through each data set
        for s in sets:

            # Set path for grid datasets
            setnum = int(s[0])

            # Get the needed photometric statistics
            nsMetrics = metrics['natsky_art']
            nsIndex = (
                (nsMetrics['Data Set'] == setnum)
            )
            pctDeg0p5 = nl_to_mccd(nsMetrics[nsIndex]['0.5 Degree Percentile'].iloc[0])
            pctDeg1p0 = nl_to_mccd(nsMetrics[nsIndex]['1.0 Degree Percentile'].iloc[0])
            pct99 = nl_to_mccd(nsMetrics[nsIndex]['99th Percentile'].iloc[0])
            pct98 = nl_to_mccd(nsMetrics[nsIndex]['98th Percentile'].iloc[0])
            pct95 = nl_to_mccd(nsMetrics[nsIndex]['95th Percentile'].iloc[0])
            pct90 = nl_to_mccd(nsMetrics[nsIndex]['90th Percentile'].iloc[0])
            pct80 = nl_to_mccd(nsMetrics[nsIndex]['80th Percentile'].iloc[0])
            pct70 = nl_to_mccd(nsMetrics[nsIndex]['70th Percentile'].iloc[0])
            pct60 = nl_to_mccd(nsMetrics[nsIndex]['60th Percentile'].iloc[0])
            pct50 = nl_to_mccd(nsMetrics[nsIndex]['Median'].iloc[0])
            pct1 = nl_to_mccd(nsMetrics[nsIndex]['1st Percentile'].iloc[0])

            # Set cell data values
            worksheet.cell(row=setnum+4, column=1 , value=dnight)           # Data night
            worksheet.cell(row=setnum+4, column=2 , value=setnum)           # Data set
            worksheet.cell(row=setnum+4, column=3 , value=pctDeg0p5)        # 0.5 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=4 , value=pctDeg1p0)        # 1.0 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 , value=pct99)            # 99th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=6 , value=pct98)            # 98th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 , value=pct95)            # 95th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=8 , value=pct90)            # 90th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 , value=pct80)            # 80th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=10, value=pct70)            # 70th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=11, value=pct60)            # 60th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=12, value=pct50)            # 50th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=13, value=pct1)             #  1st Percentile (micro-Candela)

            # Set some cell number/date formats
            worksheet.cell(row=setnum+4, column=3 ).number_format = '0.00'  # 0.5 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=4 ).number_format = '0.00'  # 1.0 Degree Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=5 ).number_format = '0.00'  # 99th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=6 ).number_format = '0.00'  # 98th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=7 ).number_format = '0.00'  # 95th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=8 ).number_format = '0.00'  # 90th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=9 ).number_format = '0.00'  # 80th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=10).number_format = '0.00'  # 70th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=11).number_format = '0.00'  # 60th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=12).number_format = '0.00'  # 50th Percentile (micro-Candela)
            worksheet.cell(row=setnum+4, column=13).number_format = '0.00'  #  1st Percentile (micro-Candela)

            # Set cell styles
            ncol = len(SHEETDATA[sheetName]['colNames'])
            for j in range(ncol):
                cell = worksheet.cell(row=setnum+4, column=j+1)
                cell.font = SHEETSTYLES['data_fields']['font']
                cell.alignment = SHEETSTYLES['data_fields']['alignment_cb']


#------------------------------------------------------------------------------#
#-------------------              Main Program              -------------------#
#------------------------------------------------------------------------------#

def generate_tables(dnight,sets,processor,centralAZ,unitName,metrics):

    # Initial status update for data set
    print(f'{PREFIX}Saving summary tables for {dnight}...')

    # Get V-band calibdata path for first data set
    calsetp1 = f"{filepath.calibdata}{dnight}/S_01/"

    # Set the output Excel file name
    excelFile = f"{filepath.tables}{dnight}.xlsx"

    # Convert central azimuth of panoramas to string
    centralAZ = int(n.floor(centralAZ))
    centralAzString = f"{centralAZ:03d}"

    # Get unit code from unit name
    unitcode = dnight[:4]

    # Get site metadata and add
    firstImage = f"{calsetp1}ib001.fit"
    siteInfo = get_site_info(firstImage)
    siteInfo['datanight'] = dnight
    siteInfo['unitname'] = unitName.replace("_"," ")
    siteInfo['unitcode'] = unitcode
    siteInfo['processor'] = processor.replace("_"," ")
    siteInfo['centralAZ'] = centralAzString
    siteInfo['numsets'] = len(sets)
    siteInfo['siteAlbedo'] = metrics['albedo']
    siteInfo['siteALR'] = metrics['alr']

    # Create the excel template
    print(f'{PREFIX}Creating empty Excel template...')
    create_excel_template(excelFile)

    # Append data to NIGHT METADATA sheet
    print(f'{PREFIX}Appending Night Metadata...')
    append_night_metadata(excelFile,siteInfo)

    # Append data to CITIES sheet
    print(f'{PREFIX}Appending top 100 cities ranked by Walkers Law...')
    append_cities(excelFile,dnight)

    # Append data to SET METADATA sheet
    print(f'{PREFIX}Appending Dataset Metadata...')
    append_set_metadata(excelFile, dnight, sets)

    # Append data to CALIBRATION sheet
    print(f'{PREFIX}Appending Image Calibration Info...')
    append_calibration(excelFile, dnight, sets)

    # Append data to EXTINCTION sheet
    print(f'{PREFIX}Appending Extinction/Zeropoint Data...')
    append_extinction(excelFile, dnight, sets)

    # Append data to IMG COORDS sheet
    print(f'{PREFIX}Appending Image Coordinates Data...')
    append_coordinates(excelFile, dnight, sets)

    # Append natural sky model params to NATSKY
    print(f'{PREFIX}Appending natural sky model parameters...')
    append_natsky_params(excelFile, dnight, sets)

    # Append data to V2 Photometry sheet
    print(f'{PREFIX}Appending V2 photometric data...')
    append_photometryV2(excelFile, dnight, sets, metrics)

    # Append data to V4 Photometry sheet
    print(f'{PREFIX}Appending V4 photometric data...')
    append_photometryV4(excelFile, dnight, sets, metrics)

    # Append data to LP sheet
    print(f'{PREFIX}Appending all-sky light pollution stats...')
    append_lp_allsky(excelFile, dnight, sets, metrics)

    # Append data to LP80 sheet
    print(f'{PREFIX}Appending ZA-80 light pollution stats...')
    append_lp_za80(excelFile, dnight, sets, metrics)

    # Append data to LP70 sheet
    print(f'{PREFIX}Appending ZA-70 light pollution stats...')
    append_lp_za70(excelFile, dnight, sets, metrics)

    # Append data to ZONES sheet
    print(f'{PREFIX}Appending Zonal light pollution stats...')
    append_zones(excelFile, dnight, sets, metrics)

    # Append data to V4 PERCENTILES ALL sheet
    print(f'{PREFIX}Appending percentile stats for all sources...')
    append_percentiles_all(excelFile, dnight, sets, metrics)

    # Append data to V4 PERCENTILES LP sheet
    print(f'{PREFIX}Appending percentile stats for artificial sources only...')
    append_percentiles_lp(excelFile, dnight, sets, metrics)