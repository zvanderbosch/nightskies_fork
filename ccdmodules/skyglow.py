#-----------------------------------------------------------------------------#
#skyglow.py
#
#NPS Night Skies Program
#
#Last updated: 2025/05/15
#
#This script computes sky luminance and illuminance
#statistics for the anthropogenic skyglow mosaic.
#
#Note: 
#
#Input:
#   (1) 
#
#Output:
#   (1) vert.xlsx - Table of vert/horiz illuminance values & stats
#
#History:
#	Zach Vanderbosch -- Created script
#
#-----------------------------------------------------------------------------#

from scipy.interpolate import make_interp_spline
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

import os
import time
import arcpy
import numpy as n
import pandas as pd
import matplotlib.pyplot as plt

# Local Source
import filepath
import printcolors as pc

# Set arcpy environment variables
arcpy.env.overwriteOutput = True
arcpy.env.rasterStatistics = "NONE"
arcpy.env.pyramid = "NONE"
arcpy.env.compression = "NONE"
arcpy.CheckOutExtension("Spatial")

# Define print staus prefix
scriptName = 'skyglow.py'
PREFIX = f'{pc.GREEN}{scriptName:19s}{pc.END}: '

# Define plotting colors for each set number
COLORS = {
    1: 'red',
    2: 'orange',
    3: 'lime',
    4: 'cornflowerblue',
    5: 'blueviolet',
    6: 'magenta',
    7: 'grey',
    8: 'black'
}

#-----------------------------------------------------------------------------#


def clear_memory(objectList):
    '''
    Function for clearing variables from memory
    '''
    for obj in objectList:
        if obj:
            del obj


def nl_to_mlux(raster):
    '''
    Unit conversion from nano-Lamberts to milli-Lux
    '''
    return (0.000000761544 * raster) / 314.159


def calc_sky_luminance(mosaicDict, zoneRaster, gridPath, results):
    '''
    Function for calculating sky luminance metrics
    '''

    # Set mosaic processing order
    mosaicOrder = ['allsky','za80','za70']

    # Iterate over each mosaic
    for i,key in enumerate(mosaicOrder):
        
        # Choose the mosaic to process
        mosaic = mosaicDict[key]

        # Calculate zonal stats
        outputTable = f"{gridPath}skyhemis{i}.dbf"
        _ = arcpy.sa.ZonalStatisticsAsTable(
            zoneRaster,"VALUE",mosaic,outputTable,"DATA","ALL"
        )

        # Extract stats from output table
        rows = arcpy.SearchCursor(outputTable)
        row = rows.next()
        results[f'skymax{i}'] = row.getValue("MAX")
        results[f'skymin{i}'] = row.getValue("MIN")
        results[f'skyave{i}'] = row.getValue("MEAN")
        clear_memory([row,rows])

    return results


def calc_zonal_sky_luminance(mosaic, zoneRaster, gridPath, results):
    '''
    Function for calculating mean sky luminance by zone
    '''

    # Calculate zonal stats
    outputTable = f"{gridPath}skyzones.dbf"
    _ = arcpy.sa.ZonalStatisticsAsTable(
        zoneRaster,"band",mosaic,outputTable,"DATA","MEAN"
    )

    # Extract stats from output table
    rows = arcpy.SearchCursor(outputTable)
    for i in range(5):
        row = rows.next()
        results[f'zoneAve{i}'] = row.getValue("MEAN")
        results[f'zoneMax{i}'] = row.getValue("COUNT")
    clear_memory([row,rows])

    return results


def calc_luminouos_emittance(mosaicDict, zoneRaster, gridPath, results):
    '''
    Function for calculating luminous emittance metrics
    '''

    # Convert from nL to mlux units
    mlux = nl_to_mlux(mosaicDict['allsky'])
    mlux80 = nl_to_mlux(mosaicDict['za80'])
    mlux70 = nl_to_mlux(mosaicDict['za70'])
    mosaicListMlux = [mlux, mlux80, mlux70]

    # Iterate over each mosaic
    for i,mosaic in enumerate(mosaicListMlux):

        # Calculate zonal stats
        outputTable = f"{gridPath}skyhemis{i}.dbf"
        _ = arcpy.sa.ZonalStatisticsAsTable(
            zoneRaster,"VALUE",mosaic,outputTable,"DATA","ALL"
        )

        # Extract stats from output table
        rows = arcpy.SearchCursor(outputTable)
        row = rows.next()
        results[f'hemis{i}'] = row.getValue("MEAN")
        results[f'totalill{i}'] = row.getValue("SUM")
        clear_memory([row,rows])

    return results


def calc_horizontal_illuminance(mosaicDict, zoneRaster, gridPath, results):
    '''
    Function for calculating horizontal illuminance metrics
    '''

    # Load in horizontal illuminance factor grid
    horizRasterFile = f"{filepath.rasters}horizillf"
    horizRaster = arcpy.sa.Raster(horizRasterFile)

    # Convert from nL to mlux units and multiply by factor grid
    mlux = nl_to_mlux(mosaicDict['allsky'])
    mlux80 = nl_to_mlux(mosaicDict['za80'])
    mlux70 = nl_to_mlux(mosaicDict['za70'])
    mosaicListHoriz = [
        mlux * horizRaster, 
        mlux80 * horizRaster, 
        mlux70 * horizRaster
    ]

    # Iterate over each mosaic
    for i,mosaic in enumerate(mosaicListHoriz):

        # Calculate zonal stats
        outputTable = f"{gridPath}skyhemis{i}.dbf"
        _ = arcpy.sa.ZonalStatisticsAsTable(
            zoneRaster,"VALUE",mosaic,outputTable,"DATA","SUM"
        )

        # Extract stats from output table
        rows = arcpy.SearchCursor(outputTable)
        row = rows.next()
        results[f'horizs{i}'] = row.getValue("SUM")
        clear_memory([row,rows])

    return results


def calc_vertical_illuminance(mosaicDict, zoneRaster, gridPath, results):
    '''
    Function for calculating vertical illuminance metrics
    '''

    # Convert skyglow from nL to mlux units
    mlux = nl_to_mlux(mosaicDict['allsky'])
    mlux80 = nl_to_mlux(mosaicDict['za80'])
    mlux70 = nl_to_mlux(mosaicDict['za70'])

    # Define array of rotation angles (0 -> 350 in 10-degree increments)
    rotationAngles = n.arange(0,360,10,dtype=int)
    numAngles = len(rotationAngles)

    # Create arrays to store illuminance values for each mosaic
    vertIllum = [
        n.zeros(numAngles),
        n.zeros(numAngles),
        n.zeros(numAngles)
    ]

    # Iterate over rotation angles in 10-degree increments
    for j,angle in enumerate(rotationAngles):

        # Status update
        print(f'{PREFIX}Calculating vertical illuminance at angle {angle}...')

        # Load in corresponding raster file
        vertRasterFile = f"{filepath.rasters}vertgrids/vertf{-angle}"
        vertRaster = arcpy.sa.Raster(vertRasterFile)

        # Calculate vertical illuminance in mlux for each zone
        mosaicListVert = [
            mlux * vertRaster,
            mlux80 * vertRaster,
            mlux70 * vertRaster
        ]

        # Iterate over each mosaic
        for i,mosaic in enumerate(mosaicListVert):

            # Calculate zonal stats
            outputTable = f"{gridPath}skyvert{i}.dbf"
            _ = arcpy.sa.ZonalStatisticsAsTable(
                zoneRaster,"VALUE",mosaic,outputTable,"DATA","SUM"
            )

            # Extract stats from output table
            rows = arcpy.SearchCursor(outputTable)
            row = rows.next()
            results[f'vert-{angle:03d}-{i}'] = row.getValue("SUM")
            vertIllum[i][j] = row.getValue("SUM")
            clear_memory([row,rows])

    # Get min/max/mean vertical illuminance values and associated azimuth values
    for i in range(len(mosaicListVert)):
        results[f'max_vlum-{i}'] = vertIllum[i].max()
        results[f'min_vlum-{i}'] = vertIllum[i].min()
        results[f'mean_vlum-{i}'] = vertIllum[i].mean()
        results[f'max_vlum_azimuth-{i}'] = rotationAngles[vertIllum[i].argmax()]
        results[f'min_vlum_azimuth-{i}'] = rotationAngles[vertIllum[i].argmin()]

    return results


def calc_sqi_histograms(zoneRaster, gridPath, clipFile80, clipFile70):
    '''
    Function to calculate Sky Quality Index (SQI) histograms
    '''

    # Clip skyglow magnitudes mosaic to 80 and 70 Zenith Angle limits
    print(f'{PREFIX}Clipping skyglow (mag) mosaic to 80 and 70 Zenith Angle limits...')
    inputRaster = f"{gridPath}skyglow/anthlightmags"
    arcpy.management.Clip(
        inputRaster,"#", "anth80mags", clipFile80,"#","ClippingGeometry"
    )
    arcpy.management.Clip(
        inputRaster,"#", "anth70mags", clipFile70,"#","ClippingGeometry"
    )

    # Create raster layers from anthropogenic mags mosaic
    print(f'{PREFIX}Generating skyglow (mag) raster layers...')
    layerName = "anthmaskedlyr"
    layerName80 = "anth80lyr"
    layerName70 = "anth70lyr"
    symbologyLayer = f"{filepath.rasters}magnitudes1.lyrx"
    arcpy.management.MakeRasterLayer(
        inputRaster,layerName,"#","#","#"
    )
    arcpy.management.MakeRasterLayer(
        "anth80mags",layerName80,"#","#","#"
    )
    arcpy.management.MakeRasterLayer(
        "anth70mags",layerName70,"#","#","#"
    )
    arcpy.management.ApplySymbologyFromLayer(
        layerName, symbologyLayer
    )
    arcpy.management.ApplySymbologyFromLayer(
        layerName80, symbologyLayer
    )
    arcpy.management.ApplySymbologyFromLayer(
        layerName70, symbologyLayer
    )

    # Iterate over each layer
    print(f'{PREFIX}Calculating SQI zonal histograms...')
    layerList = [layerName, layerName80, layerName70]
    for layer in layerList:

        # Determine output table name
        if "80" in layer:
            outputTable = f"{gridPath}sqitbl80.dbf"
        elif "70" in layer:
            outputTable = f"{gridPath}sqitbl70.dbf"
        else:
            outputTable = f"{gridPath}sqitbl.dbf"

        # Generate zonal histogram
        arcpy.sa.ZonalHistogram(
            zoneRaster, "VALUE", layer, outputTable, ""
        )

        # Delete the layer
        arcpy.management.Delete(layer,"")


def save_illuminance_data(metrics, dnight, sets, filter):
    '''
    Function to save vertical & horizontal illuminance data to vert.xlsx
    '''

    # Print status update
    print(f'{PREFIX}Saving illuminance data to spreadsheet...')

    # Define excel filename and sheet names
    excelFile = f"{filepath.calibdata}{dnight}/vert.xlsx"
    excelSheets = ["Allsky_Artificial","ZA80_Artificial","ZA70_Artificial"]

    # Get number of data sets and define column names
    Nsets = len(sets)
    excelHeaders = [' '] + [f'Set-{i+1}' for i in range(Nsets)]
    excelVertColumns = ['Azimuth'] + ['Vert Illum (mlux)']*Nsets
    excelStatRows = [
        'Min Vert (mlux)',
        'Max Vert (mlux)',
        'Mean Vert (mlux)',
        'Min Vert Azimuth',
        'Max Vert Azimuth',
        'Horiz Illum (mlux)'
    ]

    # Define some cell formatting styles
    headerFont = Font(
        name="Calibri", 
        size=11, 
        bold=True, 
        color="000000"
    )
    headerColor = PatternFill(
        fill_type='solid',
        start_color='CCFFCC', # Green
        end_color='CCFFCC'
    )
    rowColor = PatternFill(
        fill_type='solid',
        start_color='F59067', # Orange
        end_color='F59067'
    )
    headerBorder = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    headerAlignment = Alignment(
        horizontal="center", 
        vertical="center",
        wrap_text=True
    )

    # Create or open existing Excel sheet
    if not os.path.isfile(excelFile):
        try: # Create file
            writer = pd.ExcelWriter(
                excelFile, 
                engine='openpyxl'
            )
        except: # In case illumall module is using the file
            time.sleep(5)
            writer = pd.ExcelWriter(
                excelFile, 
                engine='openpyxl'
            )
    else:
        try: # Append to file
            writer = pd.ExcelWriter(
                excelFile, 
                engine='openpyxl', 
                if_sheet_exists='overlay', 
                mode='a'
            )
        except: # In case illumall module is using the file
            time.sleep(5)
            writer = pd.ExcelWriter(
                excelFile, 
                engine='openpyxl', 
                if_sheet_exists='overlay', 
                mode='a'
            )

    # Create an openpyxl workbook object
    workbook = writer.book

    # Add sheets and set column/row header formats
    for sheet in excelSheets:
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet)
            worksheet = writer.sheets[sheet]
            for i, value in enumerate(excelHeaders,1):
                cell = worksheet.cell(row=1, column=i)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = headerColor
                cell.alignment = headerAlignment
                colLetter = get_column_letter(i)
                worksheet.column_dimensions[colLetter].width = 12
            for i, value in enumerate(excelStatRows,1):
                cell = worksheet.cell(row=i+1, column=1)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = rowColor
                cell.alignment = headerAlignment
            for i, value in enumerate(excelVertColumns,1):
                cell = worksheet.cell(row=10, column=i)
                cell.value = value
                cell.font = headerFont
                cell.border = headerBorder
                cell.fill = headerColor
                cell.alignment = headerAlignment
            
            # Set width of first column separately
            worksheet.column_dimensions['A'].width = 20


    # Iterate over each sheet
    for i,sheet in enumerate(excelSheets):

        # Grab the worksheet
        worksheet = writer.sheets[sheet]

        # Iterate over each data set
        for s in sets:

            # Convert set number to integer
            setnum = int(s[0])

            # Get metrics associated with the set/filter
            setIndex = (
                (metrics['dataset'] == setnum) &
                (metrics['filter'] == filter)
            )
            setMetrics = metrics.loc[setIndex]

            # Add statistics values and format cells
            worksheet.cell(row=2, column=setnum+1, value=setMetrics[f'min_vlum-{i}'].iloc[0])
            worksheet.cell(row=3, column=setnum+1, value=setMetrics[f'max_vlum-{i}'].iloc[0])
            worksheet.cell(row=4, column=setnum+1, value=setMetrics[f'mean_vlum-{i}'].iloc[0])
            worksheet.cell(row=5, column=setnum+1, value=setMetrics[f'min_vlum_azimuth-{i}'].iloc[0])
            worksheet.cell(row=6, column=setnum+1, value=setMetrics[f'max_vlum_azimuth-{i}'].iloc[0])
            worksheet.cell(row=7, column=setnum+1, value=setMetrics[f'horizs{i}'].iloc[0])
            worksheet.cell(row=2, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=3, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=4, column=setnum+1).number_format = '0.00000'
            worksheet.cell(row=5, column=setnum+1).number_format = '0'
            worksheet.cell(row=6, column=setnum+1).number_format = '0'
            worksheet.cell(row=7, column=setnum+1).number_format = '0.00000'

            # Get vertical illuminance columns for the given sheet
            vertCols = [c for c in metrics.columns if 'vert-' in c and int(c[-1]) == i]

            # Get azimuth and vertical illuminance
            vertValues = setMetrics[vertCols].values[0]
            azValues = [int(x.split("-")[1]) for x in vertCols]

            # Add azimuth and vertical illuminance values to table
            for j,(az,vert) in enumerate(zip(azValues,vertValues)):
                
                # Add azimuth values for first set only
                if setnum == 1:
                    worksheet.cell(row=j+11, column=1, value=az)
                    worksheet.cell(row=j+11, column=1).number_format = '0'

                # Add vertical illuminance values
                worksheet.cell(row=j+11, column=setnum+1, value=vert)
                worksheet.cell(row=j+11, column=setnum+1).number_format = '0.00000'
    
    # Close the writer
    writer.close()


#-----------------------------------------------------------------------------#
# Main Program

def calculate_statistics(dnight,sets,filter):
    '''
    Main program for computing sky luminance and illuminance
    statistics from only anthropogenic sources. 
    '''

    # Filter paths
    F = {'V':'', 'B':'B/'}

    # Set ArcGIS working directories
    scratchsetp = f"{filepath.rasters}scratch_metrics/"
    arcpy.env.workspace = scratchsetp
    arcpy.env.scratchWorkspace = scratchsetp

    # Define paths to a few needed files
    zoneFile = f'{filepath.rasters}shapefiles/allbands.shp'
    za80File = f'{filepath.rasters}shapefiles/80zamaskf.shp'
    za70File = f'{filepath.rasters}shapefiles/70zamaskf.shp'
    maskRasterFile = f"{filepath.griddata}{dnight}/mask/maskd.tif"

    # Load in the mask raster
    maskRaster = arcpy.sa.Raster(maskRasterFile)

    # Create figures to plot horizontal and vertical illuminances
    figAll = plt.figure('Horizon',figsize=(10,6))
    figZ80 = plt.figure('ZA 80',figsize=(10,6))
    figZ70 = plt.figure('ZA 70',figsize=(10,6))
    axAll = figAll.add_subplot(111)
    axZ80 = figZ80.add_subplot(111)
    axZ70 = figZ70.add_subplot(111)
    axList = [axAll,axZ80,axZ70]

    # Loop through each data set
    skyglowOutput = []
    for s in sets:

        # Set path for grid datasets
        setnum = int(s[0])
        gridsetp = f"{filepath.griddata}{dnight}/S_{setnum:02d}/{F[filter]}"

        # Initial status update for data set
        print(f'{PREFIX}Processing {dnight} Set-{setnum} {filter}-band...')

        # Load in anthropogenic skyglow mosaic in nano-Lamberts [nL]
        anthRaster = arcpy.sa.Raster(f"{gridsetp}skyglow/anthlightnl")

        # Clip mosaic to 80 and 70 zenith-angle limits
        print(f'{PREFIX}Clipping skyglow (nL) mosaic to 80 and 70 Zenith Angle limits...')
        arcpy.management.Clip(
            anthRaster,"#",f"anth80{setnum}",za80File,"#","ClippingGeometry"
        )
        arcpy.management.Clip(
            anthRaster,"#",f"anth70{setnum}",za70File,"#","ClippingGeometry"
        )
        anthRaster80 = arcpy.sa.Raster(f"anth80{setnum}")
        anthRaster70 = arcpy.sa.Raster(f"anth70{setnum}")

        # Initialize result and mosaic dicts
        resultDict = {}
        mosaicDict = {
            'allsky': anthRaster,
            'za80': anthRaster80,
            'za70': anthRaster70
        }

        # Calculate sky luminance metrics
        print(f'{PREFIX}Calculating anthropogenic sky luminance...')
        resultDict = calc_sky_luminance(mosaicDict, maskRaster, gridsetp, resultDict)


        # Calculate mean sky luminance by zone
        print(f'{PREFIX}Calculating mean anthropogenic sky luminance by zone...')
        resultDict = calc_zonal_sky_luminance(anthRaster, zoneFile, gridsetp, resultDict)


        # Calculate zonal stats for each mosaic
        print(f'{PREFIX}Calculating all-sky anthropogenic luminous emittance...')
        resultDict = calc_luminouos_emittance(mosaicDict, maskRaster, gridsetp, resultDict)


        # Calculate horizontal illuminance
        print(f'{PREFIX}Calculating anthropogenic horizontal illuminance...')
        resultDict = calc_horizontal_illuminance(mosaicDict, maskRaster, gridsetp, resultDict)


        # Calculate vertical illuminance
        print(f'{PREFIX}Calculating anthropogenic vertical illuminance...')
        resultDict = calc_vertical_illuminance(mosaicDict, maskRaster, gridsetp, resultDict)


        # Calculate zonal histograms
        print(f'{PREFIX}Calculating anthropogenic SQI zonal histograms...')
        calc_sqi_histograms(maskRaster, gridsetp, za80File, za70File)

        # Generate dataframe entry for given dataset
        skyglowEntry = pd.DataFrame(
            {
                'datanight': dnight,
                'dataset': setnum,
                'filter': filter,
                **resultDict
            },
            index = [setnum-1]
        )
        skyglowOutput.append(skyglowEntry)


        ''' Add horizontal/vertical illuminance data to figures '''

        # Get resultDict keys associated with vertical illuminance
        vertKeys = [k for k in resultDict.keys() if 'vert' in k]

        # Iterate over each mask (horizon=0, ZA80=1, ZA70=2)
        for i in range(3):

            # Get the plotting axis
            ax = axList[i]

            # Get horizontal illuminance value
            horizKey = f'horizs{i}'
            horizValue = resultDict[horizKey]

            # Get rotation angles and vertical illuminance values
            vertKeysMask = [k for k in vertKeys if k[-1] == f"{i}"]
            vertAnglesMask = [float(k.split("-")[1]) for k in vertKeysMask]
            vertValuesMask = [resultDict[k] for k in vertKeysMask]

            # Plot horizontal illuminance
            ax.axhline(horizValue, ls=':', lw=2, c=COLORS[setnum], label=f'Set-{setnum} H')

            # Plot smoothed vertical illuminance
            splineInterp = make_interp_spline(vertAnglesMask, vertValuesMask)
            xsmooth = n.linspace(0.,350.,1000)
            ysmooth = splineInterp(xsmooth)
            ax.plot(xsmooth, ysmooth, ls='-', lw=2, c=COLORS[setnum], label=f'Set-{setnum} V')

            # Add plot title and axis labels and set 
            if setnum == 1:

                # Tick params
                ax.minorticks_on()
                ax.tick_params(which='both', top=True, right=True, direction='in', labelsize=13)

                # Grid lines
                ax.set_axisbelow(True)
                ax.grid(ls='-', lw=0.75, c='w')

                # Axis labels and titles
                ax.set_xlabel('Azimuth (deg)',fontsize=14)
                ax.set_ylabel('Illumination (mlux)',fontsize=14)
                if i == 0:
                    ax.set_title('Horizontal and Vertical Illuminance from Anthropogenic Light to Horizon',fontsize=15)
                elif i == 1:
                    ax.set_title('Horizontal and Vertical Illuminance from Anthropogenic Light to ZA 80',fontsize=15)
                elif i == 2:
                    ax.set_title('Horizontal and Vertical Illuminance from Anthropogenic Light to ZA 70',fontsize=15)

    # Set XY axis limits and add legend
    for ax in axList:
        ax.set_facecolor('gainsboro')
        ax.set_xlim(0,350)
        ax.set_ylim(0,1.1*ax.get_ylim()[1])
        ax.legend(
            loc='center left', 
            bbox_to_anchor=(1.01,0.5),
            handlelength=1.5,
            fontsize=11, 
            ncol=1,
            facecolor='gainsboro'
        )

    # Save figures
    print(f"{PREFIX}Saving illuminance figures...")
    figAll.savefig(
        f"{filepath.calibdata}{dnight}/illuminance_horizon.png",
        dpi=200, bbox_inches='tight'
    )
    figZ80.savefig(
        f"{filepath.calibdata}{dnight}/illuminance_za80.png",
        dpi=200, bbox_inches='tight'
    )
    figZ70.savefig(
        f"{filepath.calibdata}{dnight}/illuminance_za70.png",
        dpi=200, bbox_inches='tight'
    )

    # Create final dataframe output
    skyglowOutput = pd.concat(skyglowOutput)

    # Save illuminance data to vert.xlsx file
    save_illuminance_data(skyglowOutput, dnight, sets, filter)

    return skyglowOutput