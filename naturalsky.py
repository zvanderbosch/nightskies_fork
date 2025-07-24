#------------------------------------------------------------------------------#
#naturalsky.py
#
#NPS Night Skies Program
#
#Last updated: 2025/04/10
#
#This script generates the combined natural sky model composed
#of galactic, zodiacal, airglow, and atmospheric diffuse light
#components. The combined model is subtracted from the observed
#sky brightness to generate the anthrogenic light mosaic.
#
#Usage:
#
# python naturalsky.py DATANIGHT DATASET FILTER 
#        [--airglowzenith, --airglowheight, --airglowext, --adlfactor, --galext, --zodext]
#
# Required Arguments
# ------------------
#   DATANIGHT : Name of data night to process (e.g. ROMO241004)
#   DATASET   : Data set to process (e.g. 1, 2, 3, etc.)
#   FILTER    : Filter to process (e.g. V or B)
#
# Optional Arguments
# ------------------
#   --airglowzenith : Zenight Airglow [nL] (Default = 20)
#   --airglowheight : Height of emitting airglow layer [km] (Default = 90)
#   --airglowext    : Airglow extinction factor (Default = 0.6)
#   --adlfactor     : Atmospheric Diffuse Light factor (Default = 1.2)
#   --galext        : Galactic light extinction factor (Default = 0.9)
#   --zodext        : Zodiacal light extinction factor (Default = 0.6)
#
# Example Usage:   python naturalsky.py ROMO241004 1 V --airglowzenith=45
#
#
#Input: 
#   (1) mask.tif
#           Terrain mask after Photoshop editing (filepath.griddata/mask)
#   (2) extinction_fit_<FILTER>.xlsx
#           Bestfit extinction parameters (filepath.calibdata)
#   (3) zenithangle
#           Raster dataset providing zenith angle across the sky (filepath.rasters)
#   (4) airmass_05
#           Raster dataset providing airmass across the sky (filepath.rasters)
#   (5) adl_05
#           Raster dataset providing atmospheric diffuse light profile (filepath.rasters)
#   (6) galtopmags
#           Galactic light model in mag/arcsec^2 (filepath.griddata/S_0#/gal)
#   (7) zodtopmags
#           Zodiacal light model in mag/arcsec^2 (filepath.griddata/S_0#/zod)
#   (8) skybrightmags
#           Observed sky brightness in mag/arcsec^2 (filepath.griddata/S_0#/median)
#   (9) blankmap.aprx
#           Blank ArcGIS map project used to generate JPEG exports (filepath.maps/blankmap)
#   (10) zenith_area.shp
#           Shapefile for regional sky brightness statistics at Zenith (filepath.rasters/shapefiles)
#
#Output:
#   (1) anthlightmags<DATASET>.lyrx
#           Layer file for Anthropogenic Light Only mosaic (filepath.griddata)
#   (2) natskymags<DATASET>.lyrx
#           Layer file for Natural Sky Model mosaic (filepath.griddata)
#   (3) maskd.tif
#           Terrain mask projected into fisheye equal area coordinate system (filepath.griddata/mask)
#   (4) data.jpg
#           Fisheye JPEG image on observed sky brightness (filepath.griddata/S_0#)
#   (5) model.jpg
#           Fisheye JPEG of natural sky model (filepath.griddata/S_0#)
#   (6) artificial.jpg
#           Fisheye JPEG of anthropogenic light only (filepath.griddata/S_0#)
#   (7) hist.jpg
#           Histogram of pixel values in anthropogenic mosaic (filepath.griddata/S_0#)
#   (8) natsky_model_fit.jpeg
#           Combined data, model, artificial, and hist image (filepath.griddata/S_0#)
#   (9) <DATANIGHT>_<DATASET>_natsky_model_fit.png
#           Copy of natsky_model_fit.jpg (filepath.graphics)
#   (10) airglownl
#           Extincted airglow model in nL units (filepath.griddata/S_0#/airglow)
#   (11) galnl
#           Extincted galactic light model in nL units (filepath.griddata/S_0#/gal)
#   (12) zodnl
#           Extincted zodiacal light model in nL units (filepath.griddata/S_0#/zod)
#   (13) skybrightnl
#           Observed sky brightness in nL units (filepath.griddata/S_0#/median)
#   (14) airglowave.dbf
#           Zonal statistics table for sky brightness measurements (filepath.calibdata/S_0#)
#   (15) natsky_model_params.xlsx
#           Excel sheet summarizing natural sky model parameters & statistics (filepath.calibdata)
#
#History:
#	Dan Duriscoe -- Created as "natskyv4.py"
#	Li-Wei Hung -- Made major modification and used object-oriented approach
#   Zach Vanderbosch -- Updated to Python3 and ArcGIS Pro
#
#------------------------------------------------------------------------------#

from astropy.io import fits
from mpl_toolkits.axes_grid1 import make_axes_locatable
from skimage.transform import downscale_local_mean
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from PIL import Image

import os
import sys
import stat
import arcpy
import shutil
import argparse
import matplotlib
import matplotlib.pyplot as plt
import numpy as n
import pandas as pd

# Add path to ccdmodules
sys.path.append('./ccdmodules')

# Local Source
import filepath  
import printcolors as pc

# Print status prefix
scriptName = 'naturalsky.py'
PREFIX = f'{pc.GREEN}{scriptName:19s}{pc.END}: '

#set arcpy environment variables
arcpy.env.overwriteOutput = True
arcpy.env.rasterStatistics = "NONE"
arcpy.env.pyramid = "NONE"
arcpy.env.compression = "NONE"
arcpy.CheckOutExtension("Spatial")


#------------------------------------------------------------------------------#
#-----------------  Define Fisheye Equal Area Coord System  -------------------#
#------------------------------------------------------------------------------#

geogcs = (
    "GEOGCS["
        "'GCS_Sphere_EMEP',"
        "DATUM['D_Sphere_EMEP',"
        "SPHEROID['Sphere_EMEP',6370000.0,0.0]],"
        "PRIMEM['Greenwich',0.0],"
        "UNIT['Degree',0.0174532925199433]"
    "]"
)
coordinateSystem = (
    "PROJCS["
        "'fisheye equal area',"
        f"{geogcs},"
        "PROJECTION['Lambert_Azimuthal_Equal_Area'],"
        "PARAMETER['False_Easting',0.0],"
        "PARAMETER['False_Northing',0.0],"
        "PARAMETER['Central_Meridian',180.0],"
        "PARAMETER['Latitude_Of_Origin',90.0],"
        "UNIT['Meter',1.0]"
    "]"
)

#------------------------------------------------------------------------------#
#-------------------           Various Functions            -------------------#
#------------------------------------------------------------------------------#

#unit conversion
def nl_to_mag(x):
    """
    Converting brightness from nL to magnitude according to Dan's script
    Note: this is inconsistent to mag_to_nL()
    """
    m = 26.3308 - 2.5 * arcpy.sa.Log10(x)
    return m
    
def mag_to_nl_dan(m):
    """
    Converting brightness from magnitude to nL according to Dan's script
    Note: this is inconsistent to nL_to_mag()
    """
    x = 34.08 * arcpy.sa.Exp(20.7233 - 0.92104 * m)
    return x
    
def mag_to_nl_liwei(m):
    """
    Converting brightness from magnitude to nL according to nl_to_mag
    """
    x = 10**((26.3308-m)/2.5)
    return x

def remove_readonly(func, path):
    '''
    Error-catching function to handle removal of read-only folders
    '''
    os.chmod(path, stat.S_IWRITE)
    func(path)

def clear_scratch(scratch_dir):
    '''
    Function to clear out all files and folders from
    the scratch directory.
    '''
    for root, dirs, files in os.walk(scratch_dir, topdown=False):
        for name in files:
            try:
                os.remove(os.path.join(root, name))
            except:
                continue
        for name in dirs:
            try:
                os.chmod(os.path.join(root, name), stat.S_IWRITE)
                os.rmdir(os.path.join(root, name))
            except:
                continue

#read in galactic model, zodiacal model, and median_filtered images
def get_panoramic_raster(dnight, set, band, raster, k=25):
    """    
    This function reads in a raster file and converts it into a Python array. 
    Only area above the horizon is preserved to enforce consistency. This 
    function does not work on full resolution mosaics due limited memory space.

    Arguments:
    dnight -- data night; i.e. 'FCNA160803'
    set -- data set; i.e. 1 
    band -- filter; either 'V' or 'B'
    raster -- input raster; either 'gal', 'zod', or 'median' but not 'fullres'
    k -- downscale factor for making the image smaller

    Returns:
    A -- a 2D Python array of shape [1800,7200]/k   
    """
    filter = {'V':"",'B':"/B"}
    path = {'gal':"/gal/galtopmags",
            'zod':"/zod/zodtopmags",
            'median':filter[band]+"/median/skybrightmags"}
            
    import arcpy
    file = filepath.griddata+dnight+"/S_0"+str(set)+path[raster]
    arcpy_raster = arcpy.sa.Raster(file)  
    A = arcpy.RasterToNumPyArray(arcpy_raster, "#", "#", "#", -9999)[:1800,:7200]
    A_small = downscale_local_mean(A,(k,k))
    return A_small

#read in downscaled models and data from the fits images
def get_downscaled_image(dnight, set, band, image):
    """    
    This function reads in a downscaled fits image. Only area above the horizon 
    is preserved.
    
    Arguments:
    dnight -- data night; i.e. 'FCNA160803'
    set -- data set; i.e. 1 
    band -- filter; either 'V' or 'B'
    image -- input image; either 'gal', 'zod', or 'median' but not 'fullres'

    Returns:
    A_small -- a 2D Python array of shape [72,288]
    """
    filter = {'V':"",'B':"/B"};
    f = {'V':'', 'B':'b'}
    path = {'gal':"/galtopmags%s.fits" %set,
            'zod':"/zodtopmags%s.fits" %set,
            'median':filter[band]+"/skybrightmags%s%s.fits"%(f[band],set)}
            
    file = filepath.griddata+dnight+path[image]
    A_small = fits.open(file,unit=False)[0].data
    return A_small
    
    
#------------------------------------------------------------------------------#
#-------------------    Sky Brightness Model Components     -------------------#
#------------------------------------------------------------------------------#

#base model 
class Model(object):

    def __init__(self, dnight, set, filter, **kwargs):
        self.parameters = {}
        self.parameter_list = self.parameters.keys()
        self.dnight = dnight                                     #data night
        self.set = set                                           #data set
        self.filter = filter                                     #filter used
        self.pixscale = kwargs.get('pixscale', 0.05)             #pixscale [deg/pix]
        self.downscale = kwargs.get('downscale', 25)             #downscale factor
        self.za_min = kwargs.get('za_min', 0.)                   #min zenith angle [deg]
        self.za_max = kwargs.get('za_max', 90.)                  #max zenith angle [deg]
        self.airglow_zenith = kwargs.get('airglow_zenith', 20.)  #zenight airglow [nL]
        self.airglow_height = kwargs.get('airglow_height', 90.)  #height of airglow emitting layer [km]
        self.airglow_ext = kwargs.get('airglow_ext', 0.6)        #airglow extinction factor
        self.adl_factor = kwargs.get('adl_factor', 1.2)          #A.D.L. factor
        self.gal_ext = kwargs.get('gal_ext', 0.9)                #galactic light extinction factor
        self.zod_ext = kwargs.get('zod_ext', 0.6)                #zodiacal light extinction factor
        self.mask = kwargs.get('mask', n.array([0,]))            #terrain mask 
        self.get_extinction_coefficient()                        #extinction coeff  
        self.get_1d_za()                                         #zenith angles 1D [deg]
        self.compute_airmass()                                   #airmass 1D [deg]
        
    def get_extinction_coefficient(self,):
        """
        This function reads in the extinction coefficient associated with the 
        data set. 
        """
        d,s,f = self.dnight, self.set, self.filter
        extinctionFile = f"{filepath.calibdata}{d}/extinction_fit_{f}.xlsx"
        extinctionData = pd.read_excel(extinctionFile)
        self.extinction = abs(extinctionData['extinction_fixedZ'].iloc[s-1])
        
    def get_1d_za(self,):
        """
        This function generates a one-dimensional zenith angles values from min 
        to max with the resolution of pixscale. The output values are centered 
        in between the sampling point boundaries.
        """
        za = n.arange(self.za_min,self.za_max,self.pixscale*self.downscale)
        self.za = za + self.pixscale*self.downscale/2
        
    def compute_airmass(self,):
        """    
        This function computes the airmass at the given zenith angles according 
        to Pickering, K. A. (2002). This model has taken the atmospheric 
        refraction into account. See the airmass summary on Wikipedia.
        """
        h = 90.-self.za  # apparent altitude [deg]
        airmass = 1./n.sin(n.deg2rad(h+244./(165+47*h**1.1)))
        self.airmass = airmass[:,n.newaxis]
        
    def get_parameters(self, parameter_list=None):
        if parameter_list is None:
            parameter_list = self.parameter_list
        return [self.parameters[k] for k in parameter_list]

    def set_parameters(self, p, parameter_list=None):
        if parameter_list is None:
            parameter_list = self.parameter_list
        assert len(p) == len(parameter_list)
        for k, v in zip(parameter_list, p):
            self.parameters[k] = v

    def image_template(self, image, title, mask=False, cmapname='NPS_mag', 
                       min=14, max=24, unit='mag'):
        
        # Set colormap
        plt.rcParams['image.cmap'] = cmapname
        matplotlib.cm.get_cmap().set_bad(color='black') #mask color

        # Set mask
        if mask:
            npmask = arcpy.RasterToNumPyArray(
                self.mask,
                nodata_to_value=1
            )
            npmask = 1 - npmask # Flip 1's and 0's
            if (npmask>0).any():
                image = n.ma.MaskedArray(image,npmask)

        # Plot the image        
        fig = plt.figure(figsize=(12,3.4))
        ax = fig.add_subplot(111)
        im = ax.imshow(image, extent=(-180,180,0,90), vmin=min, vmax=max) 
        ax.set_xticks(n.arange(-180, 181, 30))
        ax.set_yticks(n.arange(0, 91, 15))
        ax.set_xlabel('Azimuth (degree)')
        ax.set_ylabel('Altitude (degree)')

        # Add colorbar
        divider = make_axes_locatable(plt.gca())
        cax = divider.append_axes("right", size="1.5%", pad="2%")
        if unit == 'mag':
            cbar = plt.colorbar(im, cax=cax,  ticks=n.arange(min,max+1))
            cbar.ax.invert_yaxis()
            labels = [str(x) for x in n.arange(min,max+1)]
            labels[0] = '<'+labels[0]
            labels[-1] = '>'+labels[-1]
            
            cbar.ax.set_yticklabels(labels)
            cbar.set_label(r'[mag / arcsec$^2$]')
        elif unit == 'nl':
            cbar = plt.colorbar(im, cax=cax)
            cbar.set_label(r'[nL]')
        elif unit == 'mask':
            cbar = plt.colorbar(im, cax=cax)
            cbar.set_label(r'Mask Value')
        
        ax.set_title(title)
        plt.tight_layout()
        plt.show(block=True)    

        
#airglow model
_AirglowModelBase = Model
class Airglow(_AirglowModelBase):

    def __init__(self, *args, **kwargs):
        # call base class constructor
        _AirglowModelBase.__init__(self, *args, **kwargs)

        self.parameter_list = self.parameters.keys()
        self.get_site_elevation()
        
    def get_site_elevation(self,):
        """
        This function reads in the elevation of the observing site from the first image in the data set.
        """
        d,s,f = self.dnight, self.set, self.filter
        F = {'V':'/','B':'/B/'}
        headerfile = f"{filepath.calibdata}{d}/S_0{s}{F[f]}ib001.fit"
        self.elevation = fits.open(headerfile)[0].header['ELEVATIO']/1000. #[km]

    def save_model_params(self,):
        """
        This function saves all input model parameters to an
        excel spreadsheet for documentation purposes.
        """

        # Excel file and sheet names
        excelFile = f"{filepath.calibdata}{self.dnight}/natsky_model_params.xlsx"
        excelSheet = "Model_Parameters"

        # Define columns to write
        excelCols = [
            "Data Set",
            "Emitting Layer Height (km)",
            "Site Elevation (m)",
            "Extinction Coefficient",
            "Zenith Airglow (nL)",
            "Airglow Extinction Constant",
            "A.D.L. Multiplier",
            "Zodiacal Extinction Constant",
            "Galactic Extinction Constant",
            "Quality Flag (0-5)",
            "Notes"
        ]

        # Create Excel file with column headers if it doesn't exist
        if not os.path.isfile(excelFile):
            with pd.ExcelWriter(excelFile, engine='openpyxl') as writer:

                # Define styles for worksheet header row
                workbook = writer.book
                headerFont = Font(
                    name="Calibri", 
                    size=11, 
                    bold=True, 
                    color="000000"
                )
                headerColor = PatternFill(
                    fill_type='solid',
                    start_color='CCFFCC',
                    end_color='CCFFCC'
                )
                headerBorder = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                headerAlignment = Alignment(
                    horizontal="center", 
                    vertical="center",
                    wrap_text=True
                )

                # Enter column headers and formatting
                workbook.create_sheet(excelSheet)
                worksheet = writer.sheets[excelSheet]
                for i, value in enumerate(excelCols,1):
                    cell = worksheet.cell(row=1, column=i)
                    cell.value = value
                    cell.font = headerFont
                    cell.border = headerBorder
                    cell.fill = headerColor
                    cell.alignment = headerAlignment
                    colLetter = get_column_letter(i)
                    if colLetter == "K":
                        worksheet.column_dimensions[colLetter].width = 90
                    else:
                        worksheet.column_dimensions[colLetter].width = 12

        # Generate input data
        dataList = [
            [self.set],
            [self.airglow_height],
            [self.elevation*1e3],
            [self.extinction],
            [self.airglow_zenith],
            [self.airglow_ext],
            [self.adl_factor],
            [self.zod_ext],
            [self.gal_ext],
            [n.nan],
            [n.nan]
        ]
        df = pd.DataFrame({k:d for k,d in zip(excelCols,dataList)})

        # Save parameters to excel sheet
        with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:
            
            # Convert DataFrame to openpyxl excel object
            df.to_excel(
                writer,
                sheet_name=excelSheet,
                startrow=self.set,
                index=None,
                header=False
            )

            # Format color and border of "Data Set" column
            worksheet = writer.sheets[excelSheet]
            fillColor = PatternFill(
                fill_type='solid',
                start_color='CCFFCC',
                end_color='CCFFCC'
            )
            thinBorder = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            worksheet[f'A{self.set+1}'].fill = fillColor
            worksheet[f'A{self.set+1}'].border = thinBorder
      
    def compute_airglow_brightness(self,a,h):
        """
        This function computes the airglow brightness according to the van Rhijn equation (Leinert et al. 1998).
        
        Arguments:
        self.za -- zenith angles [deg]
        self.elevation -- site elevation [km]
        a -- zenith airglow brightness [any linear brightness unit]
        h -- height of the emitting layer above sea level [km]
        
        Returns:
        flux -- Airglow values with the same unit as the input argument a
        """
        R = 6378+self.elevation #Earth's equatorial radius + site elevation [km]
        H = h-self.elevation   #Height of emitting layer above the observer [km]
        zaRaster = arcpy.sa.Raster(f"{filepath.rasters}zenithangle")
        sinZAraster = arcpy.sa.Sin(zaRaster * n.pi/180)
        flux = a / (1.-(R*sinZAraster/(R+H))**2)**0.5  #airglow
        return flux     

    def get_input_model(self,):
        """
        This function compute the airglow brightness [nL] at the given zenith 
        angles [deg] with default parameters.
        """
        return self.compute_airglow_brightness(20,90) #[nL]
        
    def compute_observed_model(self,):
        """
        This function computes the airglow brightness [nL] with the current
        model parameters. 
        """
        # Get airglow params
        a = self.airglow_zenith
        h = self.airglow_height
        e = self.airglow_ext

        # Load in airmass raster
        airmass = arcpy.sa.Raster(f"{filepath.rasters}airmass_05")

        # Compute observed airglow model
        airglow_nl = self.compute_airglow_brightness(a,h)  #[nL]
        airglow_mag = nl_to_mag(airglow_nl)                #[mag]
        extinction_total = e * self.extinction * airmass   #[mag]
        airglow_obs_mag = extinction_total + airglow_mag   #[mag]
        airglow_obs_nl = mag_to_nl_dan(airglow_obs_mag)    #[nL]

        return airglow_obs_nl
    
    def save_observed_model(self,):
        """
        Function to save observed Galactic model in nL units to
        a raster dataset.
        """
        d,s = self.dnight,self.set
        airglownl = self.compute_observed_model()
        airglownlPath = f"{filepath.griddata}{d}/S_0{s}/airglow/airglownl"
        airglownl.save(airglownlPath)

    def show_input_model(self,):
        """
        show the input model with default parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.get_input_model(),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Airglow_input_model", 
            cmapname='Spectral_r',
            min=n.min(img), 
            max=n.max(img),
            unit='nl'
        )
        
    def show_observed_model(self,):
        """
        show the observed model with the current parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.compute_observed_model(),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Airglow_observed_model", 
            cmapname='Spectral_r',
            min=n.min(img), 
            max=n.max(img),
            unit='nl'
        )
                

#atmospheric diffused light model
_ADLModelBase = Model
class ADL(_ADLModelBase):

    def __init__(self, *args, **kwargs):
        # call base class constructor
        _ADLModelBase.__init__(self, *args, **kwargs)

        self.parameter_list = self.parameters.keys()
        self.get_input_model()

    def get_input_model(self,):
        """    
        This function gets the atmospheric diffused light [nL] profile as a 
        function of zenith angle [deg]. The ADL is read and interpolated from 
        Dan Duriscoe's raster file called 'adl_05'. See Duriscoe 2013 and Kwon 
        et al. 2004. 
        """
        ADLraster = arcpy.sa.Raster(f"{filepath.rasters}adl_05")
        self.input_model = ADLraster
        
    def compute_observed_model(self,):
        """
        This function computes the atmospheric diffused light [nL] scaled to the
        factor a at the given zenith angles. 
        """
        return self.adl_factor * self.input_model
        
    def show_input_model(self,):
        """
        show the input model with default parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.input_model,
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "ADL_input_model", 
            cmapname='Spectral_r',
            min=n.min(img), 
            max=n.max(img),
            unit='nl'
        )
        
    def show_observed_model(self,):
        """
        show the observed model image with current parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.compute_observed_model(),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "ADL_observed_model", 
            cmapname='Spectral_r',
            min=n.min(img), 
            max=n.max(img),
            unit='nl'
        )
        

#galactic model
_GalacticModelBase = Model 
class Galactic(_GalacticModelBase):

    def __init__(self, *args, **kwargs):
        # call base class constructor
        _GalacticModelBase.__init__(self, *args, **kwargs)

        self.parameter_list = self.parameters.keys()
        self.get_input_model()
        
    def get_input_model(self,):
        """
        This function reads in the galactic light model [mag] specific for this
        set of data. 
        """
        d, s = self.dnight, self.set
            
        galPath = f"{filepath.griddata}{d}/S_0{s}/gal/galtopmags"
        galRaster = arcpy.sa.Raster(galPath)
        self.input_model = galRaster
        
    def compute_observed_model(self,unit='nl'):
        """
        This function computes the observed brightness model of galactic light 
        [mag] with the current parameters. 
        """
        airmass = arcpy.sa.Raster(f"{filepath.rasters}airmass_05")
        extinction_total = self.gal_ext * self.extinction * airmass
        if unit=='mag':
            return self.input_model + extinction_total
        else:
            return mag_to_nl_dan(self.input_model + extinction_total)
        
    def save_observed_model(self,):
        """
        Function to save observed Galactic model in nL units to
        a raster dataset.
        """
        d,s = self.dnight,self.set
        galnl = self.compute_observed_model(unit='nl')
        galnlPath = f"{filepath.griddata}{d}/S_0{s}/gal/galnl"
        galnl.save(galnlPath)
          
    def show_input_model(self,):
        """
        show the input model with default parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.input_model,
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Galactic_light_input_model"
        )

    def show_observed_model(self,):
        """
        show the observed model with the current parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.compute_observed_model(unit='mag'),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Galactic_light_obsersved_model", 
            mask=True
        )

                            
#zodiacal light model
_ZodiacalModelBase = Model
class Zodiacal(_ZodiacalModelBase):

    def __init__(self, *args, **kwargs):
        # call base class constructor
        _ZodiacalModelBase.__init__(self, *args, **kwargs)

        self.parameter_list = self.parameters.keys()
        self.get_input_model()
    
    def get_input_model(self,):
        """
        This function reads in the zodiacal light model [mag] specific for this
        set of data. 
        """
        d, s = self.dnight, self.set
        # self.input_model = get_downscaled_image(d, s, f, 'zod')

        zodPath = f"{filepath.griddata}{d}/S_0{s}/zod/zodtopmags"
        zodRaster = arcpy.sa.Raster(zodPath)
        self.input_model = zodRaster
        
    def compute_observed_model(self, unit='nl'):
        """
        This function computes the observed brightness model of zodiacal light 
        [mag] with the current parameters. 
        """
        airmass = arcpy.sa.Raster(f"{filepath.rasters}airmass_05")
        extinction_total = self.zod_ext * self.extinction * airmass
        if unit=='mag':
            return self.input_model + extinction_total
        else:
            return mag_to_nl_dan(self.input_model + extinction_total)
        
    def save_observed_model(self,):
        """
        Function to save observed Galactic model in nL units to
        a raster dataset.
        """
        d,s = self.dnight,self.set
        zodnl = self.compute_observed_model(unit='nl')
        zodnlPath = f"{filepath.griddata}{d}/S_0{s}/zod/zodnl"
        zodnl.save(zodnlPath)
        
    def show_input_model(self,):
        """
        show the input model with default parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.input_model,
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Zodiacal_light_input_model"
        )

    def show_observed_model(self,):
        """
        show the observed model with the current parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.compute_observed_model(unit='mag'),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Zodiacal_light_obsersved_model", 
            mask=True
        )


#Observed Sky Brightness
_SkyBrightObservedBase = Model
class Skybright(_SkyBrightObservedBase):

    def __init__(self, paths, *args, **kwargs):
        # call base class constructor
        _SkyBrightObservedBase.__init__(self, *args, **kwargs)

        self.paths = paths
        self.get_observed_data()
    
    def get_observed_data(self,):
        """
        This function reads in the observed median-filtered 
        sky brightness mosaic and applies the terrain mask.
        """

        # Get the necessary paths
        masksetp = self.paths['mask']
        mediansetp = self.paths['median']

        # Load terrain mask and compute stats for arcpy.sa.Con operations
        maskDataset = f"{masksetp}maskd.tif"
        mask = arcpy.sa.Raster(maskDataset)

         # Load and re-project the median mosaic
        print(f"{PREFIX}Re-projecting median mosaic to Fisheye equal area...")
        skyRaster = f"{mediansetp}skybrightmags"
        skybright = arcpy.sa.Raster(skyRaster)
        skybrightnl = mag_to_nl_dan(skybright)
        arcpy.management.ProjectRaster(
            skybrightnl, 
            "skybrightf", 
            coordinateSystem, 
            "BILINEAR",
            "5558.8"
        )

        # Apply terrain mask to median mosaic
        print(f"{PREFIX}Applying terrain mask to median mosaic...")
        skybrightnlf = arcpy.sa.Con(
            mask, "skybrightf", 0, "VALUE = 0"
        )

        # Save nL version to disk
        skybrightnl.save(f"{mediansetp}skybrightnl")

        # Save masked skybright mosaic to class
        self.masked_mosaic = skybrightnlf

        # Clear some memory
        del mask,skybright,skybrightnl,skybrightnlf

    def save_to_jpeg(self,):

        print(f"{PREFIX}Saving observed sky brightness to JPEG...")

        # Get needed paths
        gridsetp = self.paths['griddata']

        # Load in black-background ArcGIS project
        blankMap = f"{filepath.maps}blankmap/blankmap.aprx"
        p = arcpy.mp.ArcGISProject(blankMap)

        # Set map scale
        mxd = p.listMaps("Layers")[0]
        lyt = p.listLayouts()[0]
        mf = lyt.listElements()[0]
        mf.camera.scale = 120000000

        # Add sky brightness layer to data frame
        layerFile = f"{gridsetp}/skybrightmags{self.set}.lyrx"
        skybrightLayer = arcpy.mp.LayerFile(layerFile)
        mxd.addLayer(skybrightLayer)

        # Save to JPEG from MapView object.
        # Mapview is only option with width/height params for exportToJPEG
        jpegFile = f"{gridsetp}/S_0{self.set}/data.jpg"
        mv = mxd.defaultView
        mv.exportToJPEG(
            jpegFile,         # output file
            1600,             # width
            1600,             # height
            resolution=96,    # Default is 96
            jpeg_quality=100, # Default is 80, highest quality = 100
        )
        

#------------------------------------------------------------------------------#
#-------------------              Terrain Mask              -------------------#
#------------------------------------------------------------------------------#

#mask        
_MaskBase = Model
class Mask(_MaskBase):

    def __init__(self, *args, **kwargs):
        # call base class constructor
        _MaskBase.__init__(self, *args, **kwargs)

        self.get_input_model()
        
    def get_input_model(self,):
        """
        This reads in the mask to be used in the fitting process. This mask only
        extends from the zenith to the horizon. Sky = 0; terrain = 1.
        """

        # Set non-zero values of mask to Null and save to new raster
        maskTiffFile = f"{filepath.griddata}{self.dnight}/mask/mask.tif"
        maskDataset = f"{filepath.griddata}{self.dnight}/mask/maskd.tif"
        if arcpy.Exists(maskDataset):
            print(f'{PREFIX}Loading terrain mask at {maskDataset}')
            self.input_model = arcpy.sa.Raster(maskDataset)
        else:
            print(f"{PREFIX}Generating terrain mask...")
            maskRaster = arcpy.sa.Raster(maskTiffFile)

            # Define projection for the mask
            arcpy.management.DefineProjection(
                maskRaster, 
                geogcs
            )

            # Set values below horizon to NoData, above horizon to 0
            outRaster = arcpy.sa.SetNull(
                maskRaster, # Input raster
                0,          # Output value if condition is FALSE
                "VALUE = 0" # Condition. If TRUE, value set to Null
            )

            # Project mask into Fisheye equal area
            arcpy.management.ProjectRaster(
                outRaster,
                maskDataset, 
                coordinateSystem, 
                "BILINEAR",
                "5558.8"
            )

            # Compute raster statistics
            print(f"{PREFIX}Computing statistics for terrain mask...")
            arcpy.management.CalculateStatistics(maskDataset,"1","1","","","")

            # Save mask to self.input_model
            self.input_model = arcpy.sa.Raster(maskDataset)
        
    def show_input_model(self,):
        """
        show the input mask
        """
        img = arcpy.RasterToNumPyArray(
            self.input_model,
            nodata_to_value=1
        )
        self.image_template(
            img, 
            "Terrain_mask", 
            cmapname='binary',
            min=0, max=1,
            unit='mask'
        )

        
#------------------------------------------------------------------------------#
#-------------------            Combined Models             -------------------#
#------------------------------------------------------------------------------# 

_AggregateModelBase = Model
class AggregateModel(_AggregateModelBase):
    """
    This class is for combining the input models 
    - "fix_param" will keep parameters fixed during the fitting process. It
      takes the form {model:[fixed params]}
    """

    def __init__(self, model_list, paths, *args, **kwargs):
        # call base class constructor
        _AggregateModelBase.__init__(self, *args, **kwargs)

        self.model_list = model_list
        self.paths = paths
        self.fix_param = kwargs.get('fix_param',{}) #{model:[fixed params]}
    
    def compute_observed_model(self, unit='nl'):
        """
        This function computes the combined brightness from all the input models  
        """

        # Get needed paths
        gridsetp = self.paths['griddata']
        masksetp = self.paths['mask']
        natskysetp = self.paths['natsky']
        
        # Combine the models in nL units
        for i,model in enumerate(self.model_list):
            modelRaster = model.compute_observed_model()
            if i == 0:
                combinedModel = modelRaster
            else:
                combinedModel += modelRaster

        # Load terrain mask
        maskDataset = f"{masksetp}maskd.tif"
        mask = arcpy.sa.Raster(maskDataset)

        # Project natural sky model into equal area fisheye
        print(f"{PREFIX}Re-projecting natural sky model to Fisheye equal area...")
        arcpy.management.ProjectRaster(
            combinedModel,
            "natskynl", 
            coordinateSystem, 
            "BILINEAR",
            "5558.8"
        )

        # Apply mask to model
        print(f"{PREFIX}Applying terrain mask to natural sky model...")
        natskynlfc = arcpy.sa.Con(
            mask, "natskynl", 0, "VALUE = 0"
        )
        naturalskymags = nl_to_mag(natskynlfc)
        naturalskymags.save(f"{natskysetp}natskymags")

        # Save layer file
        print(f"{PREFIX}Creating natural sky model layer file...")
        layerName = f"{self.dnight}_{self.set}_natsky"
        layerFile = f"{gridsetp}/natskymags{self.set}.lyrx"
        symbologyLayer = filepath.rasters+'magnitudes.lyrx'
        arcpy.management.MakeRasterLayer(f"{natskysetp}natskymags", layerName)
        arcpy.management.ApplySymbologyFromLayer(layerName, symbologyLayer)
        arcpy.management.SaveToLayerFile(layerName, layerFile, "ABSOLUTE")

        # Clear some memory
        del mask,combinedModel

        # Return un-masked combined model for plotting purposes
        if unit == 'mag':
            return naturalskymags
        else:
            return natskynlfc
               
    def save_to_jpeg(self,):

        print(f"{PREFIX}Saving natural sky model to JPEG...")

        # Get needed paths
        gridsetp = self.paths['griddata']

        # Load in black-background ArcGIS project
        blankMap = f"{filepath.maps}blankmap/blankmap.aprx"
        p = arcpy.mp.ArcGISProject(blankMap)

        # Set map scale
        mxd = p.listMaps("Layers")[0]
        lyt = p.listLayouts()[0]
        mf = lyt.listElements()[0]
        mf.camera.scale = 120000000

        # Add natural sky layer to data frame
        layerFile = f"{gridsetp}/natskymags{self.set}.lyrx"
        natskyLayer = arcpy.mp.LayerFile(layerFile)
        mxd.addLayer(natskyLayer)

        # Save to JPEG from MapView object.
        # Mapview is only option with width/height params for exportToJPEG
        jpegFile = f"{gridsetp}/S_0{self.set}/model.jpg"
        mv = mxd.defaultView
        mv.exportToJPEG(
            jpegFile,         # output file
            1600,             # width
            1600,             # height
            resolution=96,    # Default is 96
            jpeg_quality=100, # Default is 80, highest quality = 100
        )

    def show_observed_model(self,):
        """
        show the observed model with the current parameters
        """
        img = arcpy.RasterToNumPyArray(
            self.compute_observed_model(unit='mag'),
            nodata_to_value=0
        )
        self.image_template(
            img, 
            "Natural_sky_model", 
            mask=False
        )


_SkyglowModel = Model
class SkyglowModel(_SkyglowModel):
    """
    This class is for combining the input models 
    - "fix_param" will keep parameters fixed during the fitting process. It
      takes the form {model:[fixed params]}
    """

    def __init__(self, skybright, natsky, paths, *args, **kwargs):
        # call base class constructor
        _SkyglowModel.__init__(self, *args, **kwargs)

        self.skybright = skybright
        self.natsky = natsky
        self.paths = paths
        self.compute_skyglow_model()

    def compute_skyglow_model(self,):

        # Get the necessary paths
        gridsetp = self.paths['griddata']
        skyglowsetp = self.paths['skyglow']

        # Subtract natsky model from median sky brightness mosaic
        print(f"{PREFIX}Calculating athropogenic skyglow model...")
        anthlightf = self.skybright - self.natsky
        anthlightf.save(f"{skyglowsetp}anthlightnl")

        # Remove negative values and convert to mags
        print(f"{PREFIX}Converting athropogenic skyglow model to magnitudes...")
        outCon = arcpy.sa.Con(
            anthlightf, 0.01, anthlightf, "VALUE <= 0.01"
        )
        anthlightmags = nl_to_mag(outCon)
        anthlightmags.save(f"{skyglowsetp}anthlightmags") 

        # Save magnitude mosaic to layer file
        print(f"{PREFIX}Saving athropogenic skyglow model to layer file...")
        layerName = f"{self.dnight}_{self.set}_skyglow"
        layerFile = f"{gridsetp}/anthlightmags{self.set}.lyrx"
        symbologyLayer = filepath.rasters+'magnitudes.lyrx'
        arcpy.management.MakeRasterLayer(f"{skyglowsetp}anthlightmags", layerName)
        arcpy.management.ApplySymbologyFromLayer(layerName, symbologyLayer)
        arcpy.management.SaveToLayerFile(layerName, layerFile, "ABSOLUTE")

        # Clear some memory
        del anthlightf,outCon,anthlightmags

    def save_to_jpeg(self,):
        """
        Save JPEG image of anthropogenic skyglow. This image is saved 
        in [nL] units rather than mags to help visualize areas of the 
        sky with negative values.
        """

        print(f"{PREFIX}Saving artificial skyglow model to JPEG...")

        # Get needed paths
        gridsetp = self.paths['griddata']
        skyglowsetp = self.paths['skyglow']
        scratchsetp = self.paths['scratch']

        # Load in the anthlighnl dataset and make scratch layer file
        layerFile = f"{scratchsetp}anthlightlyr.lyrx"
        symbologyLayer = filepath.rasters+'lightsub.lyrx'
        anthRaster = arcpy.sa.Raster(f"{skyglowsetp}anthlightnl")
        arcpy.management.MakeRasterLayer(anthRaster, "anthlightlyr")
        arcpy.management.ApplySymbologyFromLayer("anthlightlyr", symbologyLayer)
        arcpy.management.SaveToLayerFile("anthlightlyr", layerFile, "ABSOLUTE")

        # Load in black-background ArcGIS project
        blankMap = f"{filepath.maps}blankmap/blankmap.aprx"
        p = arcpy.mp.ArcGISProject(blankMap)

        # Set map scale
        mxd = p.listMaps("Layers")[0]
        lyt = p.listLayouts()[0]
        mf = lyt.listElements()[0]
        mf.camera.scale = 120000000

        # Add natural sky layer to data frame
        skyglowLayer = arcpy.mp.LayerFile(layerFile)
        mxd.addLayer(skyglowLayer)

        # Save to JPEG from MapView object.
        # MapView is only option with width/height params for exportToJPEG
        jpegFile = f"{gridsetp}/S_0{self.set}/artificial.jpg"
        mv = mxd.defaultView
        mv.exportToJPEG(
            jpegFile,         # output file
            1600,             # width
            1600,             # height
            resolution=96,    # Default is 96
            jpeg_quality=100, # Default is 80, highest quality = 100
        )

    
#------------------------------------------------------------------------------#
#-------------------            Mosaic Analysis             -------------------#
#------------------------------------------------------------------------------# 

_MosaicAnalysis = Model
class MosaicAnalysis(_MosaicAnalysis):

    def __init__(self, mosaic_list, paths, *args, **kwargs):

        _MosaicAnalysis.__init__(self, *args, **kwargs)
        self.mosaic_list = mosaic_list
        self.paths = paths
        stats,histarray = self.compute_zonal_stats()
        self.stats = stats
        self.histarray = histarray

    def compute_zonal_stats(self,):

        # Get path names
        skyglowsetp = self.paths['skyglow']

        # Load zenith shapefile
        inZoneData = f"{filepath.rasters}shapefiles/zenith_area.shp"

        # Iterate through each mosaic
        stat_entries = []
        for i,mosaicName in enumerate(self.mosaic_list):

            print(f"{PREFIX}Performing analysis of {mosaicName} mosaic...")

            # Determine which mosaic to analyze
            if mosaicName == 'median':
                inRaster = arcpy.sa.Raster(f"skybrightf") # Masked
            if mosaicName == 'skyglow':
                inRaster = arcpy.sa.Raster(f"{skyglowsetp}anthlightnl") # Masked

            # Execute ZonalStatisticsAsTable
            zoneField = "NAME"
            outTable = f"{filepath.calibdata}{self.dnight}/S_0{self.set}/airglowave.dbf"
            _ = arcpy.sa.ZonalStatisticsAsTable(
                inZoneData, zoneField, inRaster, outTable, "DATA", "ALL"
            )

            # Get min, max, and mean zonal stats
            rows = arcpy.SearchCursor(outTable)
            row = rows.next()
            skyave = row.getValue("MEAN")
            if row:
                del row
            if rows:
                del rows

            # Convert mosaic to array and calculate all-sky stats
            gridArray = arcpy.RasterToNumPyArray(
                inRaster, 
                nodata_to_value=-9999
            )

            # Flatten array and remove NoData values, and sort array
            maskedArray = n.ma.masked_equal(gridArray.flatten(), -9999)
            dataArray = maskedArray.compressed()
            numcells = len(dataArray)

            # Get array to return for histogram plot
            if mosaicName == 'skyglow':
                histArray = dataArray

            # Calculate min, max, mean, and median all-sky luminance
            arrmax = n.max(dataArray)
            arrmean = n.mean(dataArray)
            arrmedian = n.median(dataArray)

            # Get all-sky ALR for artificial skyglow mosaic
            if mosaicName == 'skyglow':
                # Replace values < 8.6 nL with zero
                clippedArray = n.ma.masked_less(dataArray, 8.6)
                filledArray = n.ma.filled(clippedArray, 0)

                # Calculate luminance ALR values
                illumConst = 0.00000000242407  # Not sure where this number comes from
                lumalr = arrmean / 78
                illref = numcells * 78 * illumConst # Reference condition all-sky luminance

                # calculate sum of clipped array in mlux and illuminance ALR
                clippedsum = n.ma.sum(filledArray)
                clippedmlux = illumConst * clippedsum
                illalr = clippedmlux / illref
            else:
                lumalr = n.nan
                illalr = n.nan
                clippedmlux = n.nan

            # Calculate number of square degrees, 0.5 and 1.0 degree pixel size percentiles
            sqdegrees = numcells * 0.05 * 0.05
            deg1frac = (1 - (1.00 / sqdegrees)) * 1e2
            deg05frac = (1 - (0.25 / sqdegrees)) * 1e2

            # Calculate percentiles
            percentiles = [0.05,1,60,70,80,90,95,98,99,deg1frac,deg05frac]
            pctStats = n.percentile(dataArray, percentiles)

            # Add stats to pandas DataFrame
            mosaic_entry = pd.DataFrame({
                'Data Set': self.set,
                'Mosaic_Name': mosaicName,
                'Allsky Average Luminance': arrmean,
                'Median': arrmedian,
                'Minimum (0.05 Percentile)': pctStats[0],
                'Maximum': arrmax,
                'Clipped at 8.6 nL Sum (mlux)': clippedmlux,
                'Lum_ALR': lumalr,
                'Illum_ALR': illalr,
                '0.5 Degree Percentile': pctStats[10],
                '1.0 Degree Percentile': pctStats[9],
                '99th Percentile': pctStats[8],
                '98th Percentile': pctStats[7],
                '95th Percentile': pctStats[6],
                '90th Percentile': pctStats[5],
                '80th Percentile': pctStats[4],
                '70th Percentile': pctStats[3],
                '60th Percentile': pctStats[2],
                '1st Percentile': pctStats[1],
                'Zenith': skyave
            }, index = [i])
            stat_entries.append(mosaic_entry)

        # Combine entries into single dataframe
        df_combined = pd.concat(stat_entries)

        return df_combined, histArray
    
    def save_model_stats(self,):
        """
        This function saves all computed sky brightness statistics
        to an Excel spreadsheet.
        """

        # Get the statistics
        stats = self.stats

        # Excel file and sheet names
        excelFile = f"{filepath.calibdata}{self.dnight}/natsky_model_params.xlsx"
        excelSheetMedian = "Sky_Brightness_All_Sources"
        excelSheetSkyglow = "Sky_Brightness_Artificial_Only"

        # Define columns to write to each Excel sheet
        excelColsMedian = [
            "Data Set",
            "Allsky Average Luminance",
            "Minimum (0.05 Percentile)",
            "Maximum",
            "0.5 Degree Percentile",
            "1.0 Degree Percentile",
            "99th Percentile",
            "98th Percentile",
            "95th Percentile",
            "90th Percentile",
            "80th Percentile",
            "70th Percentile",
            "60th Percentile",
            "Median",
            "1st Percentile",
            "Zenith"
        ]
        excelColsSkyGlow = [
            "Data Set",
            "Allsky Average Luminance",
            "Clipped at 8.6 nL Sum (mlux)",
            "Maximum",
            "0.5 Degree Percentile",
            "1.0 Degree Percentile",
            "99th Percentile",
            "98th Percentile",
            "95th Percentile",
            "90th Percentile",
            "80th Percentile",
            "70th Percentile",
            "60th Percentile",
            "Median",
            "1st Percentile",
            "Zenith"
        ]

        # Define some cell formatting styles
        headerFont = Font(
            name="Calibri", 
            size=11, 
            bold=True, 
            color="000000"
        )
        headerColor = PatternFill(
            fill_type='solid',
            start_color='CCFFCC',
            end_color='CCFFCC'
        )
        headerBorder = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        headerAlignment = Alignment(
            horizontal="center", 
            vertical="center",
            wrap_text=True
        )

        # Add Excel sheets with column headers if it doesn't exist
        with pd.ExcelWriter(excelFile, engine='openpyxl', mode='a') as writer:

            # Create an openpyxl workbook object
            workbook = writer.book

            # Add sheets and set column formats
            if excelSheetMedian not in workbook.sheetnames:
                workbook.create_sheet(excelSheetMedian)
                worksheet = writer.sheets[excelSheetMedian]
                for i, value in enumerate(excelColsMedian,1):
                    cell = worksheet.cell(row=1, column=i)
                    cell.value = value
                    cell.font = headerFont
                    cell.border = headerBorder
                    cell.fill = headerColor
                    cell.alignment = headerAlignment
                    colLetter = get_column_letter(i)
                    worksheet.column_dimensions[colLetter].width = 12
            if excelSheetSkyglow not in workbook.sheetnames:
                workbook.create_sheet(excelSheetSkyglow)
                worksheet = writer.sheets[excelSheetSkyglow]
                for i, value in enumerate(excelColsSkyGlow,1):
                    cell = worksheet.cell(row=1, column=i)
                    cell.value = value
                    cell.font = headerFont
                    cell.border = headerBorder
                    cell.fill = headerColor
                    cell.alignment = headerAlignment
                    colLetter = get_column_letter(i)
                    worksheet.column_dimensions[colLetter].width = 12

        # Save statistics to excel sheets
        for mosaicName in self.mosaic_list:

            if mosaicName == 'median':
                subsetStats = stats.loc[stats['Mosaic_Name'] == mosaicName][excelColsMedian]
                subsetSheet = excelSheetMedian
            elif mosaicName == 'skyglow':
                subsetStats = stats.loc[stats['Mosaic_Name'] == mosaicName][excelColsSkyGlow]
                subsetSheet = excelSheetSkyglow

            # Save parameters to excel sheet
            with pd.ExcelWriter(excelFile, engine='openpyxl', if_sheet_exists='overlay', mode='a') as writer:
                
                # Convert DataFrame to openpyxl excel object
                subsetStats.to_excel(
                    writer,
                    sheet_name=subsetSheet,
                    startrow=self.set,
                    index=None,
                    header=False
                )
                Ncol = len(subsetStats.columns.values)

                # Format color and border of "Data Set" column
                worksheet = writer.sheets[subsetSheet]
                worksheet[f'A{self.set+1}'].fill = headerColor
                worksheet[f'A{self.set+1}'].border = headerBorder

                # Format number decimal places
                for i in range(2,Ncol+1):
                    colLetter = get_column_letter(i)
                    cellID = f'{colLetter}{self.set+1}'
                    worksheet[cellID].number_format = "0.00"

    def save_summary_figure(self,):
        """
        Function that saves a summary figure displaying the images
        of the observed data, the natural sky model, and the 
        artifical sky brightness, along with a histogram of 
        luminance values in nL from the artificial sky brightness
        mosaic and some of the statistical brightness metrics.
        """

        print(f"{PREFIX}Generating summary figures...")

        # Get needed paths
        gridsetp = self.paths['griddata']
        graphicssetp = self.paths['mapgraphics']

        # Get the statistics
        stats = self.stats
        histarray = self.histarray

        #########################################################
        ## Generate the histogram

        # Generate the histogram figure
        fig = plt.figure(figsize=(24,6))
        ax = fig.add_subplot(111)

        # Figure out what bins to use
        skyglowPct80 = stats[stats['Mosaic_Name'] == 'skyglow']['80th Percentile'].iloc[0]
        if skyglowPct80 < 200:
            xlow = -49.5
            xupp = 199.5
        elif skyglowPct80 < 1000:
            xlow = -249.5
            xupp = 999.5
        elif skyglowPct80 < 5000:
            xlow = -1249.5
            xupp = 4999.5
        else:
            xlow = -2449.5
            xupp = 9999.5

        # Plot histogram
        ax.hist(
            histarray, bins=249, range=(xlow,xupp),
            histtype='bar',facecolor='b',edgecolor='k'
        )

        # Add grid lines along X-axis
        ax.grid(visible=True, which='major', axis='x', color='r', lw=3, ls='--')

        # Define some font styles for text
        font1 = {'family':'monospace', 'weight':'bold'   ,'size': 12}
        font2 = {'family':'monospace', 'weight':'normal' ,'size': 11}
        font3 = {'family':'monospace', 'weight':'bold'   ,'size': 15}

        # Add stats as text
        # xloc = -49
        xloc = 0.01
        ymax = ax.get_ylim()[1]
        skyglowIdx = (stats.Mosaic_Name == 'skyglow')
        ax.text(xloc, 0.95, 
            f"{'Mean allsky (nL)':19s}{stats['Allsky Average Luminance'][skyglowIdx].iloc[0]:>7.2f}", 
            fontdict=font1, transform=ax.transAxes)
        ax.text(xloc, 0.90, 
            f"{'Median allsky (nL)':19s}{stats['Median'][skyglowIdx].iloc[0]:>7.2f}", 
            fontdict=font1, transform=ax.transAxes)
        ax.text(xloc, 0.85, 
            f"{'Mean zenith (nL)':19s}{stats['Zenith'][skyglowIdx].iloc[0]:>7.2f}", 
            fontdict=font1, transform=ax.transAxes)
        ax.text(xloc, 0.78, 
            f"Parameters:", 
            fontdict=font1, transform=ax.transAxes)
        ax.text(xloc, 0.72, 
            f"{'Emit Lyr Ht (km)':19s}{self.airglow_height:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.67, 
            f"{'Ext Coeff':19s}{self.extinction:>7.3f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.62, 
            f"{'Zenith Airglow (nL)':19s}{self.airglow_zenith:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.57, 
            f"{'A.G. Const':19s}{self.airglow_ext:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.52, 
            f"{'Zod Const':19s}{self.zod_ext:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.47, 
            f"{'Gal Const':19s}{self.gal_ext:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.42, 
            f"{'A.D.L. Mult':19s}{self.adl_factor:>7.2f}", 
            fontdict=font2, transform=ax.transAxes)
        ax.text(xloc, 0.32, 
            f"{'LUM ALR':14s}{stats['Lum_ALR'][skyglowIdx].iloc[0]:>5.2f}", 
            fontdict=font3, transform=ax.transAxes)
        ax.text(xloc, 0.25, 
            f"{'CLIP ILL ALR':14s}{stats['Illum_ALR'][skyglowIdx].iloc[0]:>5.2f}",
            fontdict=font3, transform=ax.transAxes)
        
        # Add lines indicating clipping limits
        ax.axvline( 8.6, linewidth=1.5, color='r', ls = ':')
        ax.axvline(-8.6, linewidth=1.5, color='r', ls = ':')
        
        # Configure tick params
        ax.tick_params(which='major',direction='out',length=6,labelsize=13)

        # Axis labels and figure title
        ax.set_xlabel('Artificial sky luminance (nL)',fontsize=16)
        ax.set_ylabel('Pixel count',fontsize=16)
        title = f"{self.dnight}/S_0{self.set} DERIVED ARTIFICIAL SKY GLOW IN NANO-LAMBERTS"
        plt.title(title, fontsize=18, fontweight='bold')

        # Axis limits
        ax.set_xlim(xlow-0.5,xupp+0.5)
        ax.set_ylim(0,ymax)

        # Save histogram figure
        figFile = f"{gridsetp}/S_0{self.set}/hist.png"
        plt.savefig(figFile, dpi=200)


        #########################################################
        ## Combine histogram and data, model, artificial images

        # Load image descriptions
        imdata = Image.open(f"{graphicssetp}data.png")
        immodel = Image.open(f"{graphicssetp}model.png")
        imskyglow = Image.open(f"{graphicssetp}skyglow.png")

        # Combine all figures into one with PIL.Image
        imgPath = f"{gridsetp}/S_0{self.set}"
        im1 = Image.open(f"{imgPath}/data.jpg")
        im2 = Image.open(f"{imgPath}/model.jpg")
        im3 = Image.open(f"{imgPath}/artificial.jpg")
        im4 = Image.open(f"{imgPath}/hist.png")
        blank_image = Image.new("RGB",(4800,2800))
        blank_image.paste(im1, (0,0))
        blank_image.paste(im2, (1600,0))
        blank_image.paste(im3, (3200,0))
        blank_image.paste(im4, (0,1600))
        blank_image.paste(imdata, (1200,1500))
        blank_image.paste(immodel, (2800,1500))
        blank_image.paste(imskyglow, (4300,1500))

        # Save to griddata and graphics directories
        imgFile1 = f"{imgPath}/natsky_model_fit.png"
        imgFile2 = f"{filepath.graphics}{self.dnight}_{self.set}_natsky_model_fit.png"
        blank_image.save(imgFile1)
        blank_image.save(imgFile2)

#------------------------------------------------------------------------------#


def main():
    """
    The main execution function.
    """

    # Command line arguments
    parser = argparse.ArgumentParser()

    # Required arguments
    parser.add_argument('dataNight', type=str, 
                        help="Name of data collection night (e.g. ROMO241004)")
    parser.add_argument('dataSet', type=int, 
                        help="Data set number (e.g. 1)")
    parser.add_argument('filterName', type=str, 
                        help="Filter name (e.g. V or B)")
    
    # Optional arguments with default values
    parser.add_argument('-a','--airglowzenith', type=float, default=20.0, 
                        help='Zenight Airglow [nL] (Default = 20)')
    parser.add_argument('-t','--airglowheight', type=float, default=90.0, 
                        help='Height of emitting airglow layer [km] (Default = 90)')
    parser.add_argument('-e','--airglowext' ,type=float, default=0.6, 
                        help='Airglow extinction factor (Default = 0.6)')
    parser.add_argument('-f','--adlfactor', type=float, default=1.2,
                        help='Atmospheric Diffuse Light factor (Default = 1.2)')
    parser.add_argument('-g','--galext', type=float, default=0.9, 
                        help='Galactic light extinction factor (Default = 0.9)')
    parser.add_argument('-z','--zodext', type=float, default=0.6, 
                        help='Zodiacal light extinction factor (Default = 0.6)')
    args = parser.parse_args()

    # Pick dataset to process
    dnight = args.dataNight  #data night
    set = args.dataSet       #data set
    filter = args.filterName #filter used

    # Set directories
    gridsetp = f"{filepath.griddata}{dnight}"
    Paths = {
        'griddata': gridsetp,
        'mask': f"{gridsetp}/mask/",
        'median': f"{gridsetp}/S_0{set}/median/",
        'natsky': f"{gridsetp}/S_0{set}/nat/",
        'skyglow': f"{gridsetp}/S_0{set}/skyglow/",
        'airglow': f"{gridsetp}/S_0{set}/airglow/",
        'scratch': f"{filepath.rasters}scratch_natsky/",
        'mapgraphics': f"{filepath.maps}graphics/"
    }

    # Create/clear natsky, skyglow, airglow, and scratch directories
    for key in ['natsky','skyglow','airglow']:
        if os.path.exists(Paths[key]):
            shutil.rmtree(Paths[key], onerror=remove_readonly)
        os.makedirs(Paths[key])
    clear_scratch(Paths['scratch'])

    # Set working directories
    arcpy.env.workspace = Paths['scratch']
    arcpy.env.scratchWorkspace = Paths['scratch']

    # Setup the input parameters/kwargs
    Pa = [dnight, set, filter]
    Pk = {
        'pixscale':0.05, #unit?
        'downscale':25, 
        'za_min':0., 
        'za_max':90.,
        'airglow_zenith':args.airglowzenith,
        'airglow_height':args.airglowheight,
        'airglow_ext':args.airglowext,
        'adl_factor':args.adlfactor,
        'gal_ext': args.galext,
        'zod_ext': args.zodext
    }

    # Compute/load component models
    K = Mask(*Pa, **Pk)        # Terrain mask
    A = Airglow(*Pa, **Pk)     # Airglow model
    D = ADL(*Pa, **Pk)         # A.D.L. model
    G = Galactic(*Pa, **Pk)    # Galactic light model
    Z = Zodiacal(*Pa, **Pk)    # Zodiacal light model
    Pk['mask'] = K.input_model

    # Save model input parameters to excel sheet
    A.save_model_params()

    # Save some models to disk in [nL] units
    print(f"{PREFIX}Saving Galactic/Zodiacal/Airglow models to disk in nL units...")
    A.save_observed_model()
    G.save_observed_model()
    Z.save_observed_model()

    # Get observed sky brightness
    S = Skybright(Paths,*Pa,**Pk)
    skybright = S.masked_mosaic
    S.save_to_jpeg()

    # Get aggregate natural sky model
    M = AggregateModel([G,Z,A,D],Paths,*Pa,**Pk)
    naturalsky = M.compute_observed_model(unit='nl')
    M.save_to_jpeg()

    # Generate anthropogenic skyglow model
    X = SkyglowModel(skybright, naturalsky, Paths, *Pa, **Pk)
    X.save_to_jpeg()

    # Perform analysis of mosaics
    Q = MosaicAnalysis(['median','skyglow'],Paths,*Pa,**Pk)
    Q.save_model_stats()
    Q.save_summary_figure()

    # Clear out scratch directory
    clear_scratch(Paths['scratch'])


# Run main during script execution
if __name__ == '__main__':
    main()

#-----------------------------------------------------------------------------#