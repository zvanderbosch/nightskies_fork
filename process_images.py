#-----------------------------------------------------------------------------#
#process_images.py
#
#NPS Night Skies Program
#
#Last updated: 2025/07/24
#
#This script uses multiprocessing to reduce, calibrate, and process the images 
#collected by the NPS Night Skies Program. This script does the following:
#	(1) Reduction --- bias, dark, flat, and linearity response correction
#	(2) Registration --- image pointing determination
#	(3) Photometric Calibration --- fitting for the zeropoint and extinction
#	(4) Star Removal --- through median filtering
#	(5) Mosaicing --- producing panoramic images of the galactic model, zodiacal
#       model, median-filtered data, and full-resolution data
#
#
#Note:  
#   Importing arcpy will alter some default function name spaces (ie, time and
#   datetime). Thus, be aware of where arcpy is imported and the potential 
#   namespace conflicts. Here, arcpy is imported through importing a local 
#   mosaic source file. 
#
#
#Input:
#   (1) filelist.xlsx
#           Input list of file folders to be processed and associated parameters
#           (filepath.processlist)
#
#
#Output:
#   (1) Calibdata folder 
#           Calibrated images, processlog.txt, and other outputs (filepath.calibdata)
#   (2) Griddata folder
#           Photometrically calibrated mosaics including the galactic model, 
#           zodiacal model, median-filtered data, and full-resolution data
#           (filepath.griddata)
#
#
#History:
#	Dan Duriscoe -- Created (named firstbatchv4vb.py originally)
#	Li-Wei Hung -- Cleaned, modified, and changed some processing methods 
#   Zach Vanderbosch -- Further cleaning and optimizing parallel processing
#
#-----------------------------------------------------------------------------#

import os
import sys
import time
import shutil
import warnings
import openpyxl
import numpy as n
import pandas as pd
import astropy.units as u
import matplotlib.pyplot as plt

from astropy.io import fits
from astropy.time import Time
from datetime import datetime as Dtime
from multiprocessing import Process, Queue

# Add path to ccdmodules
sys.path.append('./ccdmodules')

# Local source (ccdmodules)
import filepath
import progressbars as pb
import printcolors as pc

# Define print status prefix
scriptName = 'process_images.py'
PREFIX = f'{pc.GREEN}{scriptName:19s}{pc.END}: '


##########################  Definitions  ######################################
def logm(m,string):
    '''To log the processing history and print status on the screen'''
    if type(string) is list:
        for s in string:
            m.append(s+'\n')
            print(s)
    else:
        m.append(string+'\n')
        print(string)    


def loghistory(message):
    '''To log the processing history and print status on the screen'''
    [history.write(line) for line in message]
    

def log_inputs(p):
    '''starting the history file and log the input files used'''
    print('Recording the reduction process in: ')
    print(filepath.calibdata+dnight+'/processlog_images.txt')
    print(' ')
    m = []
    logm(m,f"{p['Dataset']} processed on {str(Dtime.now())} by {p['Processor']}")
    logm(m,'')
    logm(m,'Input files:')
    logm(m,'Reducing V-band data? ' + p['V_band'])
    logm(m,'Reducing B-band data? ' + p['B_band'])
    logm(m,'vflat = ' + p['Flat_V'])
    logm(m,'bflat = ' + p['Flat_B'])
    logm(m,'curve = ' + p['Curve'])
    logm(m,'')
    logm(m,'____________________ Processing history: _____________________')
    logm(m,'')
    loghistory(m)
   

def update_progressbar(x,y,t=0):
    '''update the progress bar plot'''
    if t == 0:          #gray out for the in-progress status 
        barax.pcolor([x+4,x+5],[y+5,y+6],[[4],],cmap='gray',vmin=0,vmax=5)
    else:               #update for the completed status
        t/=60           #[min]
        if t < 10: texty = '%.1f' %t
        else: texty = '%.0f' %t
        if t < 6: c = 'k'
        else: c = 'w'
        Z[y+5,x] = t    #record the time [min] in a master array
        barax.pcolor([x+4,x+5],[y+5,y+6],[[t],],cmap='Greens',vmin=0,vmax=10)
        barax.text(x+4.5, y+5.5, texty, color=c, horizontalalignment='center',
                   verticalalignment='center', size='medium')
    plt.pause(0.05)     #draw the new data and run the GUI's event loop


def check_zeropoint(zp,dnight,dset):

    # The default zeropoints (mag) for each camera
    zpDefaults = {
        'IMG1': 14.67,
        'IMG2': 14.79,
        'IMG3': 14.10,
        'SBIG1': 14.79,
        'ML2': 14.68,
        'ML3': 14.78,
        'ML4': 14.62,
        'ML5': 14.71,
        'ML6': 14.75,
        'ML7': 14.93 
    }

    # Get camera name from FITS header
    firstImage = f"{filepath.rawdata}{dnight}/{dset}/ib001.fit"
    H = fits.getheader(firstImage, ext=0)
    camera = H['INSTRUME'].split(",")[0].strip().replace(" ","")

    # Compare provided and default zeropoints
    if camera in zpDefaults.keys():
        zpDefault = zpDefaults[camera]
        if not n.isclose(zp, zpDefault, atol=1e-03):
            print(f'{PREFIX}{pc.RED}WARNING{pc.END}! - Provided zeropoint ({zp:.3f}) and default zeropoint ')
            print(f'{PREFIX}for the {camera} Camera ({zpDefault:.3f}) differ by more than 0.001 mag.')
            zpcheck = input(f'{PREFIX}Continue with data processing using the provided zeropoint? (Y/N): ')
            if zpcheck.strip() not in ['Y','y']:
                sys.exit(1)


def reduce_images(*args):
    '''Basic image reduction'''
    update_progressbar(0,i)
    t1 = time.time()
    import ccdmodules.reduce as R
    if 'V' in args[2]:
        R.reducev(args[0],args[1],args[2]['V'],args[3])
    if 'B' in args[2]:
        R.reduceb(args[0],args[1],args[2]['B'],args[3])
    t2 = time.time()
    update_progressbar(0,i,t2-t1)


def register_coord(*args):
    '''Registering coordinates of each image by matching up star positions'''
    update_progressbar(1,i)
    t1 = time.time()
    m = ['Images are normally registered using full frames.','',]
    import ccdmodules.register as RE
    for filter in args[2]:
        cropped_fn, failed_fn = RE.matchstars(args[0],args[1],filter)
        logm(m,'These %s images were registered using central 200 pix:'%filter)
        logm(m,cropped_fn)
        logm(m,'')
        logm(m,'These %s images failed to be registered:'%filter)
        logm(m,failed_fn)
        logm(m,'')
    loghistory(m)
    t2 = time.time()
    update_progressbar(1,i,t2-t1)


def pointing_error(*args):
    '''Calculating the pointing error'''
    t1 = time.time()
    import ccdmodules.pointing as PO
    PO.pointing_err(*args[:-1])
    t2 = time.time()
    args[-1].put(t2-t1)


def fit_zeropoint(*args):
    '''Fitting for the extinction coefficient and zeropoint'''
    t1 = time.time()
    m = []
    import ccdmodules.extinction as EX
    for filter in args[2]:
        arg = list(args[:-1])
        arg[2] = filter
        bestfitDF = EX.extinction(*arg)

        # Log and print out fit results
        logm(m,'%s-band best fit zeropoint and extinction: ' %filter)
        columnNames = bestfitDF.columns
        bestfitp = bestfitDF.values
        nrow,ncol = bestfitp.shape
        setString = f"{'':23s}"
        for j in range(nrow):
            setString += f"Set-{j+1:<8d}"
        logm(m,setString)
        for i in range(ncol-1):
            rowString = f"{columnNames[i+1]:23s}"
            for j in range(nrow):
                if columnNames[i+1] in ['img_solved','num_star_used','num_star_rejected']:
                    rowString += f"{bestfitp[j,i+1]:<12.0f}"
                elif columnNames[i+1] in ['exptime']:
                    rowString += f"{bestfitp[j,i+1]:<12.1f}"
                else:
                    rowString += f"{bestfitp[j,i+1]:<12.3f}"
            logm(m,rowString)

    t2 = time.time()
    args[-1].put(t2-t1)
    args[-1].put(m)


def apply_filter(*args):
    '''Apply ~1 degree (in diameter) median filter to the images'''
    t1 = time.time()
    import ccdmodules.medianfilter as MF
    for filter in args[2]:
        MF.filter(args[0],args[1],filter)
    t2 = time.time()
    args[-1].put(t2-t1)


def compute_coord(*args):
    '''Computig the galactic and ecliptic coordinates of the images'''
    t1 = time.time()
    import ccdmodules.coordinates as CO
    CO.galactic_ecliptic_coords(*args[:-1])
    t2 = time.time()
    args[-1].put(t2-t1)


def mosaic_galactic(*args):
    '''Creates the mosaic of the galactic model'''
    t1 = time.time()
    import ccdmodules.galactic as GA
    GA.mosaic(*args[:-1])
    t2 = time.time()
    args[-1].put(t2-t1)


def mosaic_zodiacal(*args):
    '''Creates the mosaic of the zodiacal model'''
    t1 = time.time()
    import ccdmodules.zodiacal as ZO
    ZO.mosaic(*args[:-1])
    t2 = time.time()
    args[-1].put(t2-t1)


def mosaic_full(*args):
    '''Creates the mosaic from the full-resolution data'''
    t1 = time.time()
    import ccdmodules.fullmosaic as FM
    for filter in args[2]:
        FM.mosaic(args[0],args[1],filter)
    t2 = time.time()
    args[-1].put(t2-t1)

        
def mosaic_median(*args):
    '''Creates the mosaic from the median-filtered data'''
    t1 = time.time()
    import ccdmodules.medianmosaic as MM
    for filter in args[2]:
        MM.mosaic(args[0],args[1],filter)
    t2 = time.time()
    args[-1].put(t2-t1)
    

def utc_to_lmt(fitsHeader):
    '''
    Function to convert UTC to LMT using site longitude
    and return both as datetime objects.
    '''

    # Get header values
    dateobs = fitsHeader['DATE-OBS']
    longitude = fitsHeader['LONGITUD']

    # Get UTC date and time
    t = Time(dateobs, format='isot', scale='utc')
    utcDT = t.datetime

    # Compute LMT time in fractional hours of the day
    dayShift = 0
    hourfracUTC = t.ymdhms.hour + t.ymdhms.minute/60 + t.ymdhms.second/3600
    hourfracLMT = hourfracUTC + longitude/15
    if hourfracLMT < 0:
        hourfracLMT += 24
        dayShift -= 1

    return utcDT, hourfracLMT


def generate_calibreport(*args):
    '''
    Function to initially generate and populate 
    the calibreport.xlsx file.

    Parameters:
    -----------
    args[0] = Data Night (e.g. ROMO241004)
    args[1] = Lists of datasets to process
    args[2] = V-band flat field file
    args[3] = Linearity Curve file
    args[4] = Processor name (e.g. J Doe)
    args[5] = Location name (e.g. Rocky Mountain NP)
    '''

    # Copy calibreport template to calibdata folder
    excelTemplate = f"{filepath.spreadsheets}calibreport_template.xlsx"
    excelFile = f"{filepath.calibdata}{args[0]}/calibreport.xlsx"
    if not os.path.exists(excelFile):
        shutil.copy(excelTemplate, excelFile)

    # Get current date/time
    tnow = Dtime.now()
    datenow = tnow.strftime("%m/%d/%Y").lstrip("0")

    # Load in bestfit extinctino params
    extFile = f"{filepath.calibdata}{args[0]}/extinction_fit_V.xlsx"
    extData = pd.read_excel(extFile)

    # Open calibreport for writing
    with pd.ExcelWriter(
        excelFile, 
        engine='openpyxl', 
        if_sheet_exists='overlay', 
        mode='a') as writer:

        # Grab the Report sheet
        worksheet = writer.sheets['Report']

        # Iterate over each dataset
        for i,s in enumerate(args[1]):

            # Convert first set number to integer
            setnum = int(s[0])

            # Load in zenith images for header info
            firstImage = f"{filepath.calibdata}{args[0]}/S_{setnum:02d}/zenith1.fit"
            lastImage = f"{filepath.calibdata}{args[0]}/S_{setnum:02d}/zenith2.fit"
            firstHeader = fits.getheader(firstImage, ext=0)
            lastHeader = fits.getheader(lastImage, ext=0)

            # Get UTC and LMT datetimes
            firstUTC, firstLMTDayfrac = utc_to_lmt(firstHeader)
            lastUTC, lastLMTDayfrac = utc_to_lmt(lastHeader)

            # Get extinction fit parameters
            zeropointFixed = extData['zeropoint_default'].iloc[i]
            zeropointFree = extData['zeropoint_free'].iloc[i]
            extCoeff = extData['extinction_fixedZ'].iloc[i]
            stdErr = extData['standard_error_fixedZ'].iloc[i]
            numStarsFit = extData['num_star_used'].iloc[i]
            platescale = extData['avg_scale'].iloc[i] * 60 # arcsec/pix
            platescaleFactor = 2.5*n.log10(platescale**2)

            # Load pointing error data
            pterrFile = f"{filepath.calibdata}{args[0]}/pointerr_{setnum}.txt"
            pterrData = n.loadtxt(pterrFile)
            pterrAvg = n.mean(pterrData[:,7])

            # Save some extra info only during first set
            if i == 0:

                # Get names of calibration images
                flatFile = f'{filepath.flats}{args[2]}'
                curveFile = f'{filepath.lincurve}{args[3]}.txt'
                biasFile = f'{filepath.calibdata}{args[0]}/S_{setnum:02d}/combias.fit'
                thermalFile = f'{filepath.calibdata}{args[0]}/S_{setnum:02d}/corthermal.fit'

                # Add general data-night values to worksheet
                worksheet.cell(row=3 , column=2, value=datenow) # Processing Date
                worksheet.cell(row=3 , column=4, value=args[4]) # Processor Name
                worksheet.cell(row=3 , column=7, value=args[0]) # Data night folder
                worksheet.cell(row=4 , column=3, value=args[5]) # Long-Format Park name
                worksheet.cell(row=5 , column=3, value=firstHeader['LOCATION'])
                worksheet.cell(row=6 , column=3, value=firstHeader['LONGITUD'])
                worksheet.cell(row=7 , column=3, value=firstHeader['LATITUDE'])
                worksheet.cell(row=8 , column=3, value=firstHeader['ELEVATIO'])
                worksheet.cell(row=9 , column=3, value=firstUTC.date())
                worksheet.cell(row=10, column=3, value=firstUTC.time())
                worksheet.cell(row=12, column=3, value=firstHeader['NOTES'])
                worksheet.cell(row=4 , column=8, value=firstHeader['INSTRUME'])
                worksheet.cell(row=5 , column=8, value=firstHeader['OBSERVER'])
                worksheet.cell(row=6 , column=8, value=firstHeader['AMTEMP_F'])
                worksheet.cell(row=7 , column=8, value=firstHeader['HUMID'])
                worksheet.cell(row=8 , column=8, value=firstHeader['WIND_MPH'])
                worksheet.cell(row=9 , column=8, value=firstHeader['CCD-TEMP'])
                worksheet.cell(row=10, column=8, value=firstHeader['EXPTIME'])
                worksheet.cell(row=15, column=3, value=flatFile)
                worksheet.cell(row=16, column=3, value=curveFile)
                worksheet.cell(row=17, column=3, value=biasFile)
                worksheet.cell(row=18, column=3, value=thermalFile)
                worksheet.cell(row=15, column=3, value=flatFile)

                # Remove existing data.jpg image if present
                imagesKeep = []
                for image in worksheet._images:
                    if image.format == 'jpeg':
                        pass
                    else:
                        imagesKeep.append(image)
                worksheet._images = imagesKeep

                # Add skybright data image
                imgFile = f"{filepath.griddata}{args[0]}/S_{setnum:02d}/data.jpg"
                img = openpyxl.drawing.image.Image(imgFile)
                img.anchor = 'K4'
                img.width = 280
                img.height = 280
                worksheet.add_image(img)

            # Add data-set specific values
            worksheet.cell(row=21+i*2, column=2 , value=f"Start {firstLMTDayfrac:.2f}")
            worksheet.cell(row=22+i*2, column=2 , value=f"End {lastLMTDayfrac:.2f}")
            worksheet.cell(row=21+i*2, column=3 , value=zeropointFixed)
            worksheet.cell(row=21+i*2, column=4 , value=zeropointFree)
            worksheet.cell(row=21+i*2, column=5 , value=extCoeff)
            worksheet.cell(row=21+i*2, column=6 , value=stdErr)
            worksheet.cell(row=21+i*2, column=7 , value=numStarsFit)
            worksheet.cell(row=21+i*2, column=8 , value=platescale)
            worksheet.cell(row=21+i*2, column=9 , value=platescaleFactor)
            worksheet.cell(row=21+i*2, column=10, value=pterrAvg)





if __name__ == '__main__':
    t1 = time.time()
    #-----------------------------------------------------------------------------#    
    print(
        '\n--------------------------------------------------------------\n\n'
        '        NPS NIGHT SKIES PROGRAM RAW IMAGE PROCESSING                '
        '\n\n--------------------------------------------------------------\n'
    )
    
    warnings.filterwarnings("ignore",".*GUI is implemented.*")
        
    #------------ Read in the processing list and initialize ---------------------#

    # Read in processing dataset list and skip rows where Process = No
    filelist = pd.read_excel(f"{filepath.processlist}filelist.xlsx")
    filelist = filelist.loc[filelist.Process == 'Yes'].reset_index(drop=True)

    # Get metadata for each dataset to be processed
    Dataset = filelist['Dataset'].values
    V_band = filelist['V_band'].values
    B_band = filelist['B_band'].values
    Flat_V = filelist['Flat_V'].values
    Flat_B = filelist['Flat_B'].values
    Curve = filelist['Curve'].values
    zeropoint = filelist['Zeropoint'].values
    Processor = filelist['Processor'].values
    location = filelist['Location'].values
    
    #Check the calibration files exist    
    for i in range(len(filelist)):
        if V_band[i] == 'Yes':
            open(filepath.flats+Flat_V[i])
        if B_band[i] == 'Yes':
            open(filepath.flats+Flat_B[i])
        open(filepath.lincurve+Curve[i]+'.txt')
    
    #Determine the number of data sets collected in each night 
    img_sets = set(['1st','2nd','3rd','4th','5th','6th','7th','8th'])
    dnight_sets = {}
    nsets = []
    for dnight in Dataset:
        n_path = filepath.rawdata + dnight + '/'
        dnight_sets[dnight] = []
        for f in os.listdir(n_path):
            if os.path.isdir(n_path+f) & (f in img_sets):
                dnight_sets[dnight].append(f)
        nsets.append(len(dnight_sets[dnight]))
        
        #Make calibration folders
        if not os.path.exists(filepath.calibdata+dnight):
            os.makedirs(filepath.calibdata+dnight)
    
    
    #Plot the progress bar template
    barfig, barax = pb.bar_images(Dataset, nsets)
    if all(Processor == 'L_Hung2'):
        barfig.canvas.manager.window.move(2755,0)  
    else:
        print('You have 5 seconds to adjust the position of the progress bar window')
        plt.pause(5) #users have 5 seconds to adjust the figure position
    
    #Progress bar array (to be filled with processing time)
    Z = n.empty((5+len(filelist),14))*n.nan
    
    
    
    #------------ Main data processing code --------------------------------------#
        
    #Looping through multiple data nights
    for i in range(len(filelist)):

        history = open(filepath.calibdata+Dataset[i]+'/processlog_images.txt', 'w')
        log_inputs(filelist.loc[i])
        
        # Generate inputs for each processing step
        Filter = []; Filterset = {}
        if V_band[i] == 'Yes': 
            Filter.append('V')
            Filterset['V'] = Flat_V[i]
        if B_band[i] == 'Yes': 
            Filter.append('B')
            Filterset['B'] = Flat_B[i]
        
        #Define input arguments for various functions
        sets = dnight_sets[Dataset[i]]
        K0 = (Dataset[i],sets,Filterset,Curve[i])
        K1 = (Dataset[i],sets,Filter,zeropoint[i])
        K2 = (Dataset[i],sets,Filter) 
        K3 = (Dataset[i],sets)
        report_args = (Dataset[i],sets,Flat_V[i],Curve[i],Processor[i],location[i])

        # Check provided zeropoint against known defaults
        check_zeropoint(zeropoint[i], Dataset[i], sets[0])

        # Status update
        print(
            f'{PREFIX}Processing the {pc.BOLD}{pc.CYAN}'
            f'{Dataset[i]}{pc.END}{pc.END} dataset'
        )

        # Set up processes for each pipeline step
        q2=Queue(); Q2=(q2,); p2=Process(target=pointing_error,args=K3+Q2)
        q3=Queue(); Q3=(q3,); p3=Process(target=fit_zeropoint,args=K1+Q3)
        q4=Queue(); Q4=(q4,); p4=Process(target=apply_filter,args=K2+Q4)
        q5=Queue(); Q5=(q5,); p5=Process(target=compute_coord,args=K3+Q5) 
        q6=Queue(); Q6=(q6,); p6=Process(target=mosaic_full,args=K2+Q6)
        q7=Queue(); Q7=(q7,); p7=Process(target=mosaic_galactic,args=K3+Q7)
        q8=Queue(); Q8=(q8,); p8=Process(target=mosaic_zodiacal,args=K3+Q8)
        q9=Queue(); Q9=(q9,); p9=Process(target=mosaic_median,args=K2+Q9)
        
        # Run processes
        reduce_images(*K0)                            #image reduction   
        register_coord(*K2)                           #pointing 
        p2.start(); update_progressbar(2,i)           #pointing error
        p3.start(); update_progressbar(3,i)           #zeropoint & extinction
        p4.start(); update_progressbar(4,i)           #median filter
        p2.join() ; update_progressbar(2,i,q2.get())
        p5.start(); update_progressbar(5,i)           #galactic & ecliptic coord
        p5.join() ; update_progressbar(5,i,q5.get())
        p3.join() ; update_progressbar(3,i,q3.get())
        p6.start(); update_progressbar(6,i)           #full mosaic
        p7.start(); update_progressbar(7,i)           #galactic mosaic
        p8.start(); update_progressbar(8,i)           #zodiacal mosaic
        p9.start(); update_progressbar(9,i)           #median mosaic
        p4.join() ; update_progressbar(4,i,q4.get())
        p6.join() ; update_progressbar(6,i,q6.get())
        p7.join() ; update_progressbar(7,i,q7.get())
        p8.join() ; update_progressbar(8,i,q8.get())
        p9.join() ; update_progressbar(9,i,q9.get())

        # Generate the calibreport file
        generate_calibreport(*report_args)
        
        #log the processing history
        q_all = [q2,q3,q4,q5,q6,q7,q8,q9]
        for q in q_all:
            while not q.empty():
                loghistory(q.get())
        
        #save the timing records for running the script
        n.savetxt(filepath.calibdata+Dataset[i]+'/processtime_images.txt', Z, fmt='%4.1f')
        barfig.savefig(filepath.calibdata+Dataset[i]+'/processtime_images.png')
        
    t2 = time.time()
    print(f'{PREFIX}Total image processing time: {(t2-t1)/60:.1f} min')
    loghistory([f'Total image processing time: {(t2-t1)/60:.1f} min',])
    history.close()
    